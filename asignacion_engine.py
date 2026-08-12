# -*- coding: utf-8 -*-
"""
Motor de asignación de mercadería porcina (capón/chancha) a bloques.

Cascada de 3 capas, en el orden de prioridad que definió Tomás (2026-07-29):
  Capa 1 (prioritaria) - VOLUMEN: prorratea la cantidad nueva (plan semanal o stock del
      día) entre bloques según el peso relativo de cada uno en el Plan histórico
      (tabla A4:G13 de las hojas Capón/Chancha). Política de faltante/sobrante: prorrateo
      simple por participación, no por orden de prioridad de bloque.
  Capa 2 (calidad primaria) - peso en capón, tipificación en chancha: reparte el volumen
      de Capa 1 en sub-categorías. Usa IPF (Iterative Proportional Fitting / RAS) para
      partir del "prior" (histórico de cada bloque) y, SI hay una señal de oferta distinta
      para este lote (ej. mezcla de proveedores del plan semanal, cada uno con su propio
      perfil histórico de peso/tipif), reconciliar el reparto para que la distribución total
      por categoría también cuadre con lo que se espera recibir. Sin señal de oferta,
      col_targets = None y la capa se reduce a un prorrateo simple por bloque (equivalente a
      lo que ya hace "Detalle de calidad por vendedor" en el Excel, solo que con el total nuevo).
  Capa 3 (calidad secundaria, tercer nivel) - tipificación en capón, peso en chancha:
      prorrateo simple (sin IPF, es la prioridad más baja) del histórico de cada bloque
      dentro de cada celda ya fijada por la Capa 2.

Todo el "aprendizaje" de esta herramienta pasa por refrescar los % históricos (hojas
Capón - Peso (%), Capón - Tipificación (%), Chancha - Tipificación (%)) con ventanas más
recientes a medida que se acumulan más semanas asignadas y auditadas - no hay modelo de
ML acá a propósito: son pocos bloques y hace falta que un humano pueda auditar cada número.
"""
import datetime
import math
import re
import openpyxl

DIAS_HABILES_ES = ['Lunes', 'Martes', 'Miércoles', 'Jueves', 'Viernes']


def dia_de_reparto(fecha_faena):
    """El reparto (cuando el cliente recibe la mercadería) es el día hábil siguiente al día
    de faena — Tomás, 2026-07-31: 'la asignación se hace el día anterior (día de faena) sobre
    todo lo faenado y en stock, lo importante es el día de reparto'. Salta fines de semana
    (ej. faena del sábado -> reparto el lunes, no el domingo)."""
    d = fecha_faena + datetime.timedelta(days=1)
    while d.weekday() > 4:
        d += datetime.timedelta(days=1)
    return d

WEIGHT_BIN_LABELS = ['< 71 kg', '71 a 95 kg', '96 a 105 kg', '106 a 115 kg', '116 a 125 kg', '>= 126 kg']
TIPIF_LABELS = ['-1', '0', '1', '1+', '1++', '2', '2+', '3', '3+', '4', 'Otro / sin clasif.']


def read_pct_sheet(ws, n_categories):
    """Lee una hoja tipo 'Capón - Peso (%)': fila 4 header, desde fila 5 una fila por vendedor.
    Devuelve {codigo_vendedor: {categoria: pct}} (excluye ceros, como ya vienen guardados)."""
    header = [ws.cell(4, c).value for c in range(2, 2 + n_categories)]
    out = {}
    r = 5
    while ws.cell(r, 1).value:
        name = ws.cell(r, 1).value
        code = re.match(r'(V\d+)', name).group(1)
        d = {}
        for j, lab in enumerate(header):
            v = ws.cell(r, 2 + j).value or 0
            if v > 0.0001:
                d[lab] = v
        out[code] = d
        r += 1
    return out


DIAS_SEMANA = ['Lunes', 'Martes', 'Miércoles', 'Jueves', 'Viernes']


def read_plan_table(ws, start_row=5):
    """Lee Vendedor/Lunes..Total en A{start_row}:G..., hasta la primera fila vacía.
    Devuelve [(codigo, nombre_completo, fila, {'Lunes':.., ..., 'Viernes':.., 'Total':..}), ...]."""
    rows = []
    r = start_row
    while ws.cell(r, 1).value:
        name = ws.cell(r, 1).value
        code = re.match(r'(V\d+)', name).group(1)
        dias = {dia: (ws.cell(r, 2 + j).value or 0.0) for j, dia in enumerate(DIAS_SEMANA)}
        dias['Total'] = ws.cell(r, 7).value or 0.0
        rows.append((code, name, r, dias))
        r += 1
    return rows


class HistoricalData:
    """Carga los insumos históricos ya calculados en
    'Promedio diario Capón y Chancha por vendedor.xlsx' (Plan por bloque + % de calidad)."""

    def __init__(self, xlsx_path=None):
        if xlsx_path is None:
            # construcción vacía — usar from_rows() para poblarla (ver app_web/motor_adapter.py,
            # que arma estas mismas estructuras desde tablas de Supabase en vez de un Excel).
            self.plan_capon, self.plan_chancha = [], []
            self.peso_pct, self.tipif_pct_capon, self.tipif_pct_chancha = {}, {}, {}
            return
        wb = openpyxl.load_workbook(xlsx_path, data_only=True)
        self.plan_capon = read_plan_table(wb['Capón'])          # [(code, name, row, {dia: promedio, 'Total':..})]
        self.plan_chancha = read_plan_table(wb['Chancha'])
        self.peso_pct = read_pct_sheet(wb['Capón - Peso (%)'], 6)
        self.tipif_pct_capon = read_pct_sheet(wb['Capón - Tipificación (%)'], 10)
        self.tipif_pct_chancha = read_pct_sheet(wb['Chancha - Tipificación (%)'], 10)

    @classmethod
    def from_rows(cls, bloque_rows, calidad_rows):
        """Arma la misma estructura que __init__(xlsx_path) pero desde filas planas (ej. de
        Supabase: porcino_historico_bloque / porcino_historico_calidad, ver app_web/db.py) en
        vez de leer el Excel. bloque_rows: [{bloque_codigo, bloque_nombre, mercaderia, dia,
        cupo}]. calidad_rows: [{bloque_codigo, mercaderia, eje('peso'|'tipif'), categoria, pct}]."""
        hist = cls(xlsx_path=None)

        def _plan_table(merc):
            por_bloque = {}
            for r in bloque_rows:
                if r['mercaderia'] != merc:
                    continue
                por_bloque.setdefault(r['bloque_codigo'], {'nombre': r['bloque_nombre'], 'dias': {}})
                por_bloque[r['bloque_codigo']]['dias'][r['dia']] = float(r['cupo'])
            out = []
            for code, d in por_bloque.items():
                dias = {dia: d['dias'].get(dia, 0.0) for dia in DIAS_SEMANA}
                dias['Total'] = d['dias'].get('Total', 0.0)
                out.append((code, d['nombre'], None, dias))
            return out

        def _pct(merc, eje):
            out = {}
            for r in calidad_rows:
                if r['mercaderia'] != merc or r['eje'] != eje:
                    continue
                out.setdefault(r['bloque_codigo'], {})[r['categoria']] = float(r['pct'])
            return out

        hist.plan_capon = _plan_table('Capón')
        hist.plan_chancha = _plan_table('Chancha')
        hist.peso_pct = _pct('Capón', 'peso')
        hist.tipif_pct_capon = _pct('Capón', 'tipif')
        hist.tipif_pct_chancha = _pct('Chancha', 'tipif')
        return hist

    def bloque_shares(self, merc, dia=None):
        """{codigo: (nombre, share)} = participación de cada bloque en el histórico del Plan.
        Si 'dia' se pasa (ej. 'Jueves'), la participación se calcula sobre el promedio histórico
        de ESE día de la semana puntual (cada bloque reparte distinto según el día — ver
        Capón!B5:F12 — no alcanza con su participación en el total semanal). Sin 'dia', usa el
        total semanal (uso: Modo Proyectado, donde no hay un día puntual todavía)."""
        rows = self.plan_capon if merc == 'Capón' else self.plan_chancha
        clave = dia if dia else 'Total'
        total = sum(r[3][clave] for r in rows)
        return {code: (name, (dias[clave] / total if total else 0.0)) for code, name, _, dias in rows}

    def cupo_bloque(self, merc, dia=None):
        """{codigo: (nombre, cupo_absoluto)} — el promedio histórico ABSOLUTO (no normalizado a
        participación) de ese día de la semana puntual (o el total semanal sin 'dia'). Es "lo que
        hay que cumplir" (Tomás, 2026-07-30): un objetivo fijo por bloque, no una cuota que se
        estira o encoge para vaciar todo el stock del día."""
        rows = self.plan_capon if merc == 'Capón' else self.plan_chancha
        clave = dia if dia else 'Total'
        return {code: (name, dias[clave]) for code, name, _, dias in rows}

    def cupo_o_cuota(self, merc, dia=None, cuota_semana=None, dias_referencia=None):
        """Igual forma que cupo_bloque ({codigo: (nombre, valor)}), pero si se pasa
        cuota_semana ({codigo: total pedido por ese bloque para TODA la semana}), lo usa como
        base — repartido entre los días de la semana según la FORMA histórica propia de ese
        bloque — en vez del promedio histórico absoluto. Tomás, 2026-08-02: "el cupo histórico
        es relativo, nos sirve como orientación... lo importante es lo que pide para la semana.
        Si no piden nada usamos el cupo histórico [como referencia]. Y el histórico nos tiene
        que servir para saber qué días de la semana demandan más o menos y en qué calidades."
        O sea: el TOTAL semanal lo manda la cuota real cuando existe; el histórico sigue
        marcando el reparto entre días (Capa 1) y las calidades (Capa 2/3, sin cambios).
        dias_referencia: si 'cuota_semana' ya viene NETA de días resueltos por fuera (ej. el
        reparto de ayer, ver asignacion_semana), pasar la lista de días que quedan por repartir
        — la fracción de cada día se recalcula SOLO entre esos días (no diluida con la forma
        histórica de un día que ya no corresponde repartir), así el total de los días restantes
        sigue sumando exacto cuota_semana. Sin este parámetro, usa el 'Total' semanal completo
        (comportamiento de siempre)."""
        base = self.cupo_bloque(merc, dia=dia)  # {code: (name, cupo_hist_absoluto_del_dia)}
        if not cuota_semana:
            return base
        rows = self.plan_capon if merc == 'Capón' else self.plan_chancha
        clave = dia if dia else 'Total'
        if dias_referencia:
            frac_dia = {code: (dias[clave] / sum(dias[d] for d in dias_referencia)
                                if sum(dias[d] for d in dias_referencia) else 0.0)
                        for code, _, _, dias in rows}
        else:
            frac_dia = {code: (dias[clave] / dias['Total'] if dias['Total'] else 0.0)
                        for code, _, _, dias in rows}
        out = {}
        for code, (name, cupo_hist) in base.items():
            cuota = cuota_semana.get(code)
            if cuota:  # pidió algo esta semana -> la cuota manda, repartida con la forma del día
                out[code] = (name, cuota * frac_dia.get(code, 0.0))
            else:  # no pidió nada -> cupo histórico, como referencia (comportamiento anterior)
                out[code] = (name, cupo_hist)
        return out

    def distribuir_cuota_dias(self, merc, cuota_semana, dias):
        """{codigo: {dia: cantidad_entera}} — reparte la cuota semanal de CADA bloque entre
        'dias' según la forma histórica PROPIA de ese bloque (qué días entrega más/menos),
        SIN caparla contra stock/faena disponible ese día. El total semanal de cada bloque da
        EXACTO su cuota (Tomás, 2026-08-02: "el total por bloque debe ser igual a la cuota que
        él pide en la semana... distribuí la cuota en los días que tiene mayor entrega").
        Si el bloque no cargó cuota (0/ausente), cae al cupo histórico absoluto de cada día
        (mismo fallback que cupo_o_cuota) — el histórico ahí SÍ define el total, a falta de
        otro dato; cuando hay cuota, el histórico solo decide el reparto entre días."""
        rows = self.plan_capon if merc == 'Capón' else self.plan_chancha
        cuota_semana = cuota_semana or {}
        out = {}
        for code, _name, _r, dias_hist in rows:
            cuota = cuota_semana.get(code)
            if cuota:
                total_hist = dias_hist['Total']
                floats = {d: (cuota * dias_hist[d] / total_hist if total_hist else cuota / len(dias)) for d in dias}
                out[code] = round_preserve_sum(floats, cuota)
            else:
                out[code] = {d: round(dias_hist[d]) for d in dias}
        return out

    def distribuir_cuota_restante(self, merc, cuota_semana, dias_todos, dias_fijos):
        """Como distribuir_cuota_dias, pero con días YA resueltos por fuera (ej. el reparto de
        mañana ya se asignó hoy en modo manual, ver ABASTO Pedidos) — Tomás, 2026-08-03: "ya no
        van a demandar ochocientos y pico de capones, debés restarle todo lo que se reparte
        mañana". dias_fijos: {dia: {codigo: cantidad_ya_asignada}}. Para cada bloque: cuota
        restante = cuota_semana − lo ya asignado en días fijos; esa cuota restante se reparte
        SOLO entre los días libres (no en dias_fijos), con la forma histórica de esos días
        libres (recalculada entre ellos, sin diluirla con el peso histórico del día ya fijado).
        Devuelve {codigo: {dia: cantidad}} — incluye tanto los días fijos (tal cual) como los
        libres (recalculados)."""
        rows = self.plan_capon if merc == 'Capón' else self.plan_chancha
        cuota_semana = cuota_semana or {}
        dias_libres = [d for d in dias_todos if d not in dias_fijos]
        out = {}
        for code, _name, _r, dias_hist in rows:
            fijo_total = sum(dias_fijos.get(d, {}).get(code, 0) for d in dias_fijos)
            resultado = {d: dias_fijos[d].get(code, 0) for d in dias_fijos if code in dias_fijos[d]}
            cuota = cuota_semana.get(code)
            if cuota:
                restante = max(cuota - fijo_total, 0)
                total_libre_hist = sum(dias_hist[d] for d in dias_libres)
                floats = {d: (restante * dias_hist[d] / total_libre_hist if total_libre_hist
                              else restante / len(dias_libres)) for d in dias_libres}
                resultado.update(round_preserve_sum(floats, restante))
            else:
                resultado.update({d: round(dias_hist[d]) for d in dias_libres})
            out[code] = resultado
        return out

    def redirigir(self, origen, destino):
        """Funde el cupo diario y participación histórica de 'origen' dentro de 'destino' —
        para cuando un bloque deja de operar y su cartera pasa a otro de forma provisoria
        (Tomás, 2026-08-02: V04 dejó de trabajar, su cartera de capón pasa a V11 "por ahora,
        después asignaremos mejor su cartera de clientes"). No borra el histórico de 'origen',
        solo deja de aparecer como bloque propio en esta corrida — si se vuelve a necesitar,
        no llamar a este método."""
        for attr in ('plan_capon', 'plan_chancha'):
            rows = getattr(self, attr)
            origen_dias = next((dias for code, _, _, dias in rows if code == origen), None)
            nuevas = [t for t in rows if t[0] != origen]
            if origen_dias:
                nuevas = [
                    (code, name, r, {k: dias[k] + origen_dias[k] for k in dias}) if code == destino else (code, name, r, dias)
                    for code, name, r, dias in nuevas
                ]
            setattr(self, attr, nuevas)
        for pct_attr in ('peso_pct', 'tipif_pct_capon', 'tipif_pct_chancha'):
            getattr(self, pct_attr).pop(origen, None)


def capa1_prorrateo(total_nuevo, shares):
    """{codigo: cantidad_objetivo} — prorratea total_nuevo según el % histórico de cada bloque.
    Uso: Modo Proyectado, donde SÍ hace falta repartir el total completo (es una cantidad que se
    va a comprar/traer, no un stock ya disponible que puede sobrar)."""
    return {code: total_nuevo * share for code, (_, share) in shares.items()}


def capa1_cupo(cupo_bloque, total_disponible):
    """Capa 1 para Modo Stock disponible: "lo que hay que cumplir es el cupo, no es necesario que
    quedemos en stock cero" (Tomás, 2026-07-30). cupo_bloque: {codigo: cupo_absoluto_del_día}.
    Si el stock disponible ALCANZA el cupo total, cada bloque recibe su cupo completo y lo que
    sobra queda sin asignar (no se fuerza a repartir todo el stock). Si el stock disponible NO
    alcanza, se prorratea la faltante proporcionalmente al cupo de cada bloque (misma política de
    faltante/sobrante ya acordada: prorrateo por volumen histórico trabajado por bloque).
    Devuelve (target_bloque_float, total_a_asignar_entero) — total_a_asignar es cuánto del stock
    disponible realmente se reparte hoy (== total_disponible si hay faltante, o el cupo total
    redondeado PARA ARRIBA si sobra — Tomás, 2026-08-06: "hay que cumplir con el cupo, y si hay
    stock disponible redondear para arriba", para no quedarse corto por un simple redondeo
    cuando hay stock de sobra para cubrir el cupo completo)."""
    total_cupo = sum(cupo_bloque.values())
    if total_cupo <= 0:
        return {code: 0.0 for code in cupo_bloque}, 0
    if total_disponible >= total_cupo:
        return dict(cupo_bloque), math.ceil(total_cupo)
    factor = total_disponible / total_cupo
    return {code: v * factor for code, v in cupo_bloque.items()}, total_disponible


def ipf_balance(seed, row_targets, col_targets, max_iter=200, tol=1e-7):
    """RAS / Iterative Proportional Fitting. seed: {row: {col: valor_prior}}.
    Ajusta 'seed' al mínimo necesario para que las filas sumen row_targets[row] y las
    columnas sumen col_targets[col], preservando lo más posible las proporciones originales.
    Si col_targets es None, no hay señal de oferta nueva: se devuelve el prorrateo simple
    de cada fila a su row_target (sin tocar columnas)."""
    rows = list(row_targets.keys())
    # key=str: las categorías pueden ser tuplas (peso, tipif) con peso=None cuando el stock
    # todavía no tiene tipificación/peso cargado (ver stock_real.filas_manuales) — comparar
    # None con str directamente rompe sorted(), por eso se ordena por su representación textual.
    cols = sorted({c for r in rows for c in seed.get(r, {})}, key=str)
    m = {r: {c: seed.get(r, {}).get(c, 0.0) for c in cols} for r in rows}

    def scale_rows():
        for r in rows:
            s = sum(m[r].values())
            if s > 1e-12:
                f = row_targets[r] / s
                for c in cols:
                    m[r][c] *= f

    scale_rows()
    if not col_targets:
        return m  # sin señal de oferta: prorrateo simple por bloque, capa 2 = histórico de cada bloque

    for _ in range(max_iter):
        col_sums = {c: sum(m[r][c] for r in rows) for c in cols}
        max_dev = max(abs(col_sums[c] - col_targets.get(c, 0.0)) for c in cols)
        if max_dev < tol:
            break
        for c in cols:
            s = col_sums[c]
            target = col_targets.get(c, 0.0)
            if s > 1e-12:
                f = target / s
                for r in rows:
                    m[r][c] *= f
        scale_rows()
    return m


def capa2_calidad_primaria(target_bloque, demand_pct, supply_profile=None):
    """target_bloque: {codigo: cantidad}. demand_pct: {codigo: {categoria: pct histórico}}.
    supply_profile: {categoria: pct} de este lote en particular (ej. derivado de la mezcla de
    proveedores del plan semanal) o None si no hay señal distinta al histórico.
    Devuelve {codigo: {categoria: cantidad}}."""
    seed = {code: {cat: target_bloque[code] * pct for cat, pct in demand_pct.get(code, {}).items()}
            for code in target_bloque}
    col_targets = None
    if supply_profile:
        total = sum(target_bloque.values())
        col_targets = {cat: pct * total for cat, pct in supply_profile.items()}
    return ipf_balance(seed, target_bloque, col_targets)


def capa3_prorrateo(capa2_matrix, secondary_pct):
    """capa2_matrix: {codigo: {cat_primaria: cantidad}}. secondary_pct: {codigo: {cat_secundaria: pct}}.
    Prorrateo simple (prioridad más baja, sin IPF) dentro de cada celda ya fijada por la capa 2.
    Devuelve {codigo: {cat_primaria: {cat_secundaria: cantidad}}}."""
    out = {}
    for code, by_primary in capa2_matrix.items():
        out[code] = {}
        sec = secondary_pct.get(code, {})
        for cat1, qty in by_primary.items():
            if qty <= 1e-9:
                continue
            out[code][cat1] = {cat2: qty * pct for cat2, pct in sec.items()}
    return out


def round_preserve_sum(values, total_int=None):
    """Método del resto mayor (Hamilton): redondea 'values' (dict clave->float) a enteros que
    suman EXACTO total_int (o round(sum(values)) si no se pasa). Nunca se "pierde" ni se
    "inventa" una cabeza por redondeo — cada unidad entera de stock queda asignada a alguien.
    Reparte primero el piso de cada valor, y las unidades que faltan para completar el total
    van a quienes tenían el resto (la parte decimal) más grande."""
    if not values:
        return {}
    if total_int is None:
        total_int = round(sum(values.values()))
    floors = {k: math.floor(v) for k, v in values.items()}
    falta = total_int - sum(floors.values())
    if falta <= 0:
        return floors
    orden = sorted(values.keys(), key=lambda k: values[k] - floors[k], reverse=True)
    out = dict(floors)
    for k in orden[:falta]:
        out[k] += 1
    return out


def enterizar_cascada(target_bloque_float, capa2_float, capa3_float, total_int):
    """Convierte la cascada completa (Capa1 float -> Capa2 float -> Capa3 float, todas con
    decimales) en cantidades ENTERAS que suman exacto en cada nivel: el total de bloques suma
    total_int, dentro de cada bloque la Capa2 suma el entero de ese bloque, y dentro de cada
    celda de Capa2 la Capa3 suma el entero de esa celda. Así cada cabeza real de stock queda
    asignada a un bloque + categoría concreta, no a un valor con decimales.
    Devuelve (bloque_int, capa2_int, capa3_int) con la misma forma que las versiones float."""
    bloque_int = round_preserve_sum(target_bloque_float, total_int)

    capa2_int = {}
    for code, total_bloque in bloque_int.items():
        capa2_int[code] = round_preserve_sum(capa2_float.get(code, {}), total_bloque)

    capa3_int = {}
    if capa3_float is not None:
        for code, by_cat1 in capa2_int.items():
            capa3_int[code] = {}
            for cat1, total_cat1 in by_cat1.items():
                if total_cat1 <= 0:
                    continue
                sub_float = capa3_float.get(code, {}).get(cat1, {})
                capa3_int[code][cat1] = round_preserve_sum(sub_float, total_cat1)
    return bloque_int, capa2_int, capa3_int


def transportation_round(float_matrix, row_targets_int, col_targets_int):
    """Redondea una matriz continua (ya balanceada por IPF, así que sus filas/columnas suman
    ~row_targets/col_targets) a números ENTEROS que suman EXACTO tanto por fila como por
    columna — requiere sum(row_targets_int) == sum(col_targets_int) (es el caso siempre acá:
    ambos son el mismo total de cabezas reales). Método: piso en cada celda, y las unidades
    que faltan para completar cada fila/columna se agregan una por una a la celda con mayor
    parte decimal entre las que todavía tienen margen en su fila y su columna (heurística
    estándar de redondeo controlado 2D — no siempre es la única solución óptima, pero es
    determinística, auditable y siempre factible cuando los totales cuadran)."""
    rows = list(row_targets_int)
    cols = list(col_targets_int)
    m = {r: {c: math.floor(float_matrix.get(r, {}).get(c, 0.0)) for c in cols} for r in rows}
    row_rem = {r: row_targets_int[r] - sum(m[r].values()) for r in rows}
    col_rem = {c: col_targets_int[c] - sum(m[r][c] for r in rows) for c in cols}

    fractional = sorted(
        ((float_matrix.get(r, {}).get(c, 0.0) - m[r][c], r, c) for r in rows for c in cols),
        key=lambda t: t[0], reverse=True)
    for _frac, r, c in fractional:
        if row_rem[r] > 0 and col_rem[c] > 0:
            m[r][c] += 1
            row_rem[r] -= 1
            col_rem[c] -= 1

    # remanente residual (empates o casos borde): completar con lo que quede de margen
    for r in rows:
        while row_rem[r] > 0:
            c = next((c for c in cols if col_rem[c] > 0), None)
            if c is None:
                break
            m[r][c] += 1
            row_rem[r] -= 1
            col_rem[c] -= 1
    return m


def reconciliar_con_stock_real(target_bloque_int, target_bloque_float, demand_pct, real_counts, bloque_shares):
    """Capa 2/3 unificada cuando SÍ hay stock real itemizado (no una proyección): reparte
    exactamente las unidades REALMENTE disponibles de cada categoría (real_counts) entre
    bloques, arrancando de la forma que da el histórico de cada bloque (demand_pct) como
    'prior' vía IPF, y redondeando con transportation_round para que no sobre ni falte ni
    una unidad real. demand_pct: {codigo: {categoria: pct}}. real_counts: {categoria: cantidad
    entera realmente disponible}. Si ninguna categoría histórica de NINGÚN bloque cubre una
    categoría que sí existe en el stock real, esa categoría se reparte por el peso relativo
    general de cada bloque (bloque_shares) en vez de perderse.
    Devuelve {codigo: {categoria: cantidad_entera}}."""
    cats_con_historial = {c for pcts in demand_pct.values() for c in pcts}
    seed = {}
    for code, target_f in target_bloque_float.items():
        pcts = demand_pct.get(code, {})
        seed[code] = {cat: target_f * pcts.get(cat, 0.0) for cat in real_counts}
    for cat in real_counts:
        if cat not in cats_con_historial:
            for code, target_f in target_bloque_float.items():
                seed[code][cat] = target_f * bloque_shares.get(code, (None, 0.0))[1]

    m = ipf_balance(seed, target_bloque_int, real_counts)
    return transportation_round(m, target_bloque_int, real_counts)


def _orden_fifo(f):
    # normaliza a date: stock real trae datetime.datetime (celda Excel), las filas sintéticas
    # de faena pendiente (stock_real.filas_manuales) traen datetime.date — mezclar ambos tipos
    # rompe la comparación de sorted().
    fecha = f['fecha_faena']
    if hasattr(fecha, 'date') and callable(fecha.date):
        fecha = fecha.date()
    return (fecha or datetime.date.min, str(f['correlativo']))


def asignar_correlativos_fifo(bloque_order, matrix_int, stock_por_categoria, orden_key=_orden_fifo):
    """Asigna correlativos REALES (no proyectados) a cada bloque, categoría por categoría,
    con desempate por orden_key (default: FIFO por fecha de faena, el más viejo primero) —
    ataca directamente el problema de deshidratación por falta de FIFO que ya está documentado
    en Origen Pampa. Para bovino se pasa un orden_key que además respeta la PRIORIDAD ya
    calculada por Origen Pampa (MUY ALTA primero) antes que los días de faena.
    stock_por_categoria: {categoria: [fila_stock, ...]} ya ordenadas o no (se ordenan acá).
    Devuelve {codigo: [fila_stock, ...]} y muta stock_por_categoria (consume lo asignado)."""
    for cat, filas in stock_por_categoria.items():
        filas.sort(key=orden_key)
    asignado = {code: [] for code in bloque_order}
    for cat in list(stock_por_categoria):
        cola = stock_por_categoria[cat]
        for code in bloque_order:
            n = matrix_int.get(code, {}).get(cat, 0)
            asignado[code].extend(cola[:n])
            del cola[:n]
    return asignado


def expandir_correlativos(bloque_names, bloque_int, capa2_int, capa3_int, primary_order, secondary_order):
    """Expande la cascada entera a una lista de "cabeza por cabeza": un renglón por unidad real
    de stock, con el correlativo (1..N) y a qué bloque/categoría le toca. Es literalmente lo que
    pidió Tomás: "que termines asignando cada correlativo en los distintos bloques" — nada de
    porcentajes ni decimales, cada fila es un animal físico y a quién se le entrega."""
    filas = []
    n = 0
    for code, (name, _share) in bloque_names.items():
        cats1 = capa2_int.get(code, {})
        for cat1 in primary_order:
            cant1 = cats1.get(cat1, 0)
            if cant1 <= 0:
                continue
            if capa3_int is None:
                for _ in range(cant1):
                    n += 1
                    filas.append({'correlativo': n, 'bloque': name, 'primaria': cat1, 'secundaria': None})
                continue
            cats2 = capa3_int.get(code, {}).get(cat1, {})
            for cat2 in secondary_order:
                cant2 = cats2.get(cat2, 0)
                for _ in range(cant2):
                    n += 1
                    filas.append({'correlativo': n, 'bloque': name, 'primaria': cat1, 'secundaria': cat2})
    return filas
