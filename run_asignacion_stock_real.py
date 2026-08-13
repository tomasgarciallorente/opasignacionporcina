# -*- coding: utf-8 -*-
"""
Asignación con stock disponible REAL (no proyectado): toma cada carcaza real de
STOCK (BD) — correlativo, kg, tipificación, fecha de faena reales — y decide a qué
bloque le toca cada una, aplicando la cascada acordada con Tomás (2026-07-29/30):
  Capa 1 - cantidad (prorrateo por volumen histórico del bloque)
  Capa 2/3 unificadas - calidad (peso+tipificación en capón, tipificación en chancha),
      reconciliadas contra lo que REALMENTE hay disponible hoy, no contra una proyección.
Desempate dentro de cada categoría: FIFO por fecha de faena (el más viejo primero) —
ataca directo el problema de deshidratación por falta de FIFO ya documentado en la empresa.

Uso:
    python run_asignacion_stock_real.py --out "Propuesta - Stock real 2026-07-29.xlsx"
"""
import argparse
import datetime
from collections import Counter, defaultdict

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

from asignacion_engine import (
    HistoricalData, capa1_cupo, capa1_prorrateo, round_preserve_sum,
    reconciliar_con_stock_real, asignar_correlativos_fifo, dia_de_reparto,
)
from stock_real import read_stock_bd

STOCK_XLSX = r'C:\Users\Gtecomercial\Dropbox\01-Abasto\03-Entregas de carnes\2B-OP Stock y Entregas Porcino 2026.xlsx'
HIST_XLSX = r'C:\Users\Gtecomercial\OneDrive\Desktop\Mente AI ORIGEN PAMPA\Análisis de entregas\Promedio diario Capón y Chancha por vendedor.xlsx'
OUT_DIR = r'C:\Users\GTECOM~1\AppData\Local\Temp\claude\C--Users-Gtecomercial-OneDrive-Desktop-Mente-AI-ORIGEN-PAMPA--claude\e3eb4ea7-112c-4af9-b68f-acca9a5b4f03\scratchpad'

HEADER_FILL = PatternFill('solid', fgColor='1F2937')
HEADER_FONT = Font(color='FFFFFF', bold=True, size=11)
VENDOR_FILL = PatternFill('solid', fgColor='334155')
VENDOR_FONT = Font(color='FFFFFF', bold=True, size=12)
TITLE_FONT = Font(bold=True, size=13)
SUB_FONT = Font(italic=True, size=10, color='555555')
WARN_FILL = PatternFill('solid', fgColor='FEF3C7')
APPROVE_FILL = PatternFill('solid', fgColor='DCFCE7')
THIN = Side(style='thin', color='CCCCCC')
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
CENTER = Alignment(horizontal='center', vertical='center')


def _asignar_generico(hist, filas, merc, dia, demand_joint_fn, cat_de_fila, cuota_semana=None,
                       dias_referencia=None, cupo_override=None, modo='cupo'):
    """cat_de_fila(fila) -> categoría (tupla (peso,tipif) para capón, o tipif solo para chancha).
    cuota_semana: {codigo: total pedido esa semana} — si un bloque pidió algo, manda eso (ver
    HistoricalData.cupo_o_cuota); si no pidió nada, cae al cupo histórico de siempre.
    dias_referencia: días que quedan por repartir cuando cuota_semana ya viene neta de días
    resueltos por fuera — ver HistoricalData.cupo_o_cuota.
    cupo_override: {codigo: cupo_absoluto} — si se pasa, IGNORA cuota_semana/dias_referencia y
    usa este cupo tal cual (uso: segunda pasada del día, ver asignacion_semana — el cupo que
    quedó pendiente después de agotar la faena vieja, para cubrir con la faena nueva del día).
    modo: 'cupo' (default, mismo comportamiento de siempre) — cubre el cupo histórico/cuota y
    deja el resto sin asignar en 'Sobrante' si el stock alcanza y sobra.
    modo='prorrateo' — Tomás, 2026-08-12 ("no me puedo quedar con stock sin asignar"): reparte
    el 100% del stock disponible en proporción a la participación histórica de cada bloque
    (capa1_prorrateo), sin dejar sobrante — usado por la app puntual, ver motor_adapter.py."""
    shares = hist.bloque_shares(merc, dia=dia)
    if modo == 'prorrateo':
        total_disponible = len(filas)
        target_f = capa1_prorrateo(total_disponible, shares)
        total_a_asignar = total_disponible
        cupo = {code: v for code, (_name, v) in shares.items()}
    else:
        if cupo_override is not None:
            cupo = dict(cupo_override)
        else:
            cupo = {code: v for code, (_name, v) in hist.cupo_o_cuota(
                merc, dia=dia, cuota_semana=cuota_semana, dias_referencia=dias_referencia).items()}
        total_disponible = len(filas)
        target_f, total_a_asignar = capa1_cupo(cupo, total_disponible)
    target_i = round_preserve_sum(target_f, total_a_asignar)

    demand_joint = demand_joint_fn()
    real_counts_full = Counter(cat_de_fila(r) for r in filas)
    col_scale = (total_a_asignar / total_disponible) if total_disponible else 0.0
    real_counts_scaled = {cat: cnt * col_scale for cat, cnt in real_counts_full.items()}
    real_counts_i = round_preserve_sum(real_counts_scaled, total_a_asignar)

    matrix = reconciliar_con_stock_real(target_i, target_f, demand_joint, real_counts_i, shares)

    stock_por_cat = defaultdict(list)
    for r in filas:
        stock_por_cat[cat_de_fila(r)].append(r)
    bloque_order = list(shares.keys())
    asignado = asignar_correlativos_fifo(bloque_order, matrix, stock_por_cat)
    sobrante = [r for filas_cat in stock_por_cat.values() for r in filas_cat]  # lo que quedó sin asignar
    return shares, target_i, matrix, asignado, total_disponible, total_a_asignar, sobrante, cupo


def asignar_capon(hist, stock_rows, dia=None, cuota_semana=None, dias_referencia=None, cupo_override=None, modo='cupo'):
    filas = [r for r in stock_rows if r['merc'] == 'Capón']

    def demand_joint_fn():
        d = {}
        for code in hist.bloque_shares('Capón', dia=dia):
            pesos = hist.peso_pct.get(code, {})
            tipifs = hist.tipif_pct_capon.get(code, {})
            d[code] = {(p, t): pv * tv for p, pv in pesos.items() for t, tv in tipifs.items()}
        return d

    return _asignar_generico(hist, filas, 'Capón', dia, demand_joint_fn, lambda r: (r['peso'], r['tipif']),
                              cuota_semana=cuota_semana, dias_referencia=dias_referencia,
                              cupo_override=cupo_override, modo=modo)


def asignar_chancha(hist, stock_rows, dia=None, cuota_semana=None, dias_referencia=None, cupo_override=None, modo='cupo'):
    filas = [r for r in stock_rows if r['merc'] == 'Chancha']
    return _asignar_generico(hist, filas, 'Chancha', dia, lambda: hist.tipif_pct_chancha, lambda r: r['tipif'],
                              dias_referencia=dias_referencia, cuota_semana=cuota_semana,
                              cupo_override=cupo_override, modo=modo)


def write_resumen(ws, merc, total_disponible, total_a_asignar, dia, shares, target_i, matrix, con_peso):
    ws['A1'] = f'Criterio de asignación — {merc} — para el reparto del {dia or "(semana)"}'
    ws['A1'].font = TITLE_FONT
    ws.merge_cells('A1:D1')
    ws['A2'] = 'ESTADO: PENDIENTE DE APROBACIÓN'
    ws['A2'].font = Font(bold=True, size=11, color='92400E')
    ws['A2'].fill = WARN_FILL
    ws.merge_cells('A2:D2')
    sobrante_n = total_disponible - total_a_asignar
    cupo_txt = (f'Se cubre el cupo completo de {dia or "la semana"} ({total_a_asignar}) y quedan '
                f'{sobrante_n} carcazas sin asignar en stock (ver solapa Sobrante).' if sobrante_n > 0
                else f'El stock disponible ({total_disponible}) no alcanza el cupo de {dia or "la semana"} '
                     f'— se prorrateó la faltante entre bloques.')
    ws['A3'] = (f'Stock disponible (faenado hasta hoy): {total_disponible} carcazas reales (STOCK (BD)). {cupo_txt} '
                f'Cantidad y calidad de cada correlativo asignado son datos reales — ver solapa Correlativos. '
                f'La asignación se decide hoy (día de faena) para el reparto del día hábil siguiente.')
    ws['A3'].font = SUB_FONT
    ws.merge_cells('A3:D3')
    ws.row_dimensions[3].height = 30
    ws['A3'].alignment = Alignment(wrap_text=True, vertical='top')
    ws['A4'] = 'Aprobado por:'
    ws['A4'].font = Font(bold=True)
    ws['B4'].fill = APPROVE_FILL
    ws['B4'].border = BORDER
    ws['C4'] = 'Fecha de aprobación:'
    ws['C4'].font = Font(bold=True)
    ws['D4'].fill = APPROVE_FILL
    ws['D4'].border = BORDER

    r = 6
    for code, (name, _share) in shares.items():
        ws.cell(r, 1, name)
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=4)
        ws.cell(r, 1).font = VENDOR_FONT
        ws.cell(r, 1).fill = VENDOR_FILL
        r += 1
        ws.cell(r, 1, 'Total asignado').border = BORDER
        ws.cell(r, 1).font = Font(bold=True)
        ws.cell(r, 2, target_i.get(code, 0)).border = BORDER
        ws.cell(r, 2).font = Font(bold=True)
        ws.cell(r, 2).alignment = CENTER
        r += 1
        cats = matrix.get(code, {})
        for cat, qty in sorted(cats.items(), key=lambda kv: -kv[1]):
            if qty <= 0:
                continue
            label = f'{cat[0]} / tipif. {cat[1]}' if con_peso else f'tipif. {cat}'
            ws.cell(r, 1, '   ▸ ' + label).border = BORDER
            ws.cell(r, 2, qty).border = BORDER
            ws.cell(r, 2).alignment = CENTER
            r += 1
        r += 1
    ws.column_dimensions['A'].width = 34
    for col in 'BCD':
        ws.column_dimensions[col].width = 16


def write_correlativos(ws, merc, asignado, shares):
    ws['A1'] = f'Correlativos reales asignados — {merc} — {datetime.date.today().isoformat()}'
    ws['A1'].font = TITLE_FONT
    ws.merge_cells('A1:G1')
    headers = ['Bloque asignado', 'Correlativo', 'Kg', 'Tipificación', 'Peso', 'Fecha faena', 'Proveedor']
    for j, h in enumerate(headers):
        c = ws.cell(3, 1 + j, h)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.alignment = CENTER
        c.border = BORDER
    r = 4
    for code, (name, _share) in shares.items():
        for f in asignado.get(code, []):
            ws.cell(r, 1, name).border = BORDER
            ws.cell(r, 2, f['correlativo']).border = BORDER
            ws.cell(r, 2).alignment = CENTER
            ws.cell(r, 3, f['kg']).border = BORDER
            ws.cell(r, 3).alignment = CENTER
            ws.cell(r, 4, f['tipif']).border = BORDER
            ws.cell(r, 4).alignment = CENTER
            ws.cell(r, 5, f['peso'] or '').border = BORDER
            fecha = f['fecha_faena']
            ws.cell(r, 6, fecha.date() if hasattr(fecha, 'date') else fecha).border = BORDER
            ws.cell(r, 6).number_format = 'dd/mm/yyyy'
            ws.cell(r, 6).alignment = CENTER
            ws.cell(r, 7, f['proveedor']).border = BORDER
            r += 1
    for col, w in [('A', 26), ('B', 12), ('C', 8), ('D', 14), ('E', 14), ('F', 14), ('G', 20)]:
        ws.column_dimensions[col].width = w
    ws.freeze_panes = 'A4'


def write_sobrante(ws, merc, sobrante):
    ws['A1'] = f'Sin asignar hoy (queda en stock) — {merc} — {datetime.date.today().isoformat()}'
    ws['A1'].font = TITLE_FONT
    ws.merge_cells('A1:E1')
    ws['A2'] = ('Carcazas que ya cubrieron el cupo de todos los bloques para hoy y quedan en stock '
                'para el próximo reparto — no hace falta vaciarlas hoy.')
    ws['A2'].font = SUB_FONT
    ws.merge_cells('A2:E2')
    headers = ['Correlativo', 'Kg', 'Tipificación', 'Peso', 'Fecha faena', 'Proveedor']
    for j, h in enumerate(headers):
        c = ws.cell(4, 1 + j, h)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.alignment = CENTER
        c.border = BORDER
    r = 5
    def _orden_sobrante(f):
        fecha = f['fecha_faena']
        if hasattr(fecha, 'date') and callable(fecha.date):
            fecha = fecha.date()
        return (fecha or datetime.date.min, str(f['correlativo']))

    for f in sorted(sobrante, key=_orden_sobrante):
        ws.cell(r, 1, f['correlativo']).border = BORDER
        ws.cell(r, 1).alignment = CENTER
        ws.cell(r, 2, f['kg']).border = BORDER
        ws.cell(r, 2).alignment = CENTER
        ws.cell(r, 3, f['tipif']).border = BORDER
        ws.cell(r, 3).alignment = CENTER
        ws.cell(r, 4, f['peso'] or '').border = BORDER
        fecha = f['fecha_faena']
        ws.cell(r, 5, fecha.date() if hasattr(fecha, 'date') else fecha).border = BORDER
        ws.cell(r, 5).number_format = 'dd/mm/yyyy'
        ws.cell(r, 5).alignment = CENTER
        ws.cell(r, 6, f['proveedor']).border = BORDER
        r += 1
    for col, w in [('A', 12), ('B', 8), ('C', 14), ('D', 14), ('E', 14), ('F', 20)]:
        ws.column_dimensions[col].width = w
    ws.freeze_panes = 'A5'


def write_resumen_dia(ws, dia, bloques_por_merc):
    """Cuadrito único: por mercadería y bloque, cupo del día / asignado / diferencia, más una
    fila de remanente sin asignar (stock que queda para el reparto siguiente, no es de ningún
    bloque en particular — a diferencia de una diferencia negativa, que sí es de un bloque
    puntual que se quedó corto). Pedido de Tomás (2026-07-30): "un cuadrito resumen de todo lo
    que se asigna en el día... y el remanente para el día siguiente, y de qué bloque es"."""
    ws['A1'] = f'Resumen — para el reparto del {dia or ""} (decidido hoy {datetime.date.today().isoformat()}, día de faena)'
    ws['A1'].font = TITLE_FONT
    ws.merge_cells('A1:E1')
    ws['A2'] = ('Cupo = lo que ese bloque necesita hoy según el histórico. Diferencia negativa = '
                'ese bloque se quedó corto hoy (no había suficiente stock de su calidad). La fila '
                '"Remanente sin asignar" es stock que sobra en general — no es de ningún bloque, '
                'queda disponible para el próximo reparto (ver solapa Sobrante de cada mercadería).')
    ws['A2'].font = SUB_FONT
    ws.merge_cells('A2:E2')
    ws.row_dimensions[2].height = 30
    ws['A2'].alignment = Alignment(wrap_text=True, vertical='top')

    headers = ['Mercadería', 'Bloque', 'Cupo del día', 'Asignado hoy', 'Diferencia']
    for j, h in enumerate(headers):
        c = ws.cell(4, 1 + j, h)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.alignment = CENTER
        c.border = BORDER

    r = 5
    total_cupo_gral, total_asig_gral = 0, 0
    for merc, (shares, cupo, target_i, disp, asig) in bloques_por_merc.items():
        merc_start = r
        subtotal_cupo, subtotal_asig = 0, 0
        # round_preserve_sum en vez de round() independiente por bloque -- mismo bug de
        # "los totales no coinciden" que se resolvió en consolidado_porcino.py/
        # build_presupuesto_estatico (Tomás, 2026-08-06). Esta tabla día-a-día ya no la usa
        # consolidado_porcino.py (reemplazada por el Resumen semanal - Bloque), pero se deja
        # consistente por si se vuelve a usar sola (ver run_asignacion.py).
        cupo_int = round_preserve_sum(cupo)
        for code, (name, _s) in shares.items():
            cupo_v = cupo_int.get(code, 0)
            asig_v = target_i.get(code, 0)
            diff = asig_v - cupo_v
            subtotal_cupo += cupo_v
            subtotal_asig += asig_v
            ws.cell(r, 1, merc).border = BORDER
            ws.cell(r, 2, name).border = BORDER
            ws.cell(r, 3, cupo_v).border = BORDER
            ws.cell(r, 3).alignment = CENTER
            ws.cell(r, 4, asig_v).border = BORDER
            ws.cell(r, 4).alignment = CENTER
            dc = ws.cell(r, 5, diff)
            dc.border = BORDER
            dc.alignment = CENTER
            if diff < 0:
                dc.font = Font(color='B91C1C', bold=True)
            r += 1
        remanente = disp - asig
        sr = r
        ws.cell(sr, 1).fill = PatternFill('solid', fgColor='E5E7EB')
        ws.cell(sr, 1).border = BORDER
        ws.cell(sr, 2, f'Subtotal {merc}').font = Font(bold=True)
        ws.cell(sr, 2).fill = PatternFill('solid', fgColor='E5E7EB')
        ws.cell(sr, 2).border = BORDER
        ws.cell(sr, 3, subtotal_cupo).font = Font(bold=True)
        ws.cell(sr, 3).alignment = CENTER
        ws.cell(sr, 3).fill = PatternFill('solid', fgColor='E5E7EB')
        ws.cell(sr, 3).border = BORDER
        ws.cell(sr, 4, subtotal_asig).font = Font(bold=True)
        ws.cell(sr, 4).alignment = CENTER
        ws.cell(sr, 4).fill = PatternFill('solid', fgColor='E5E7EB')
        ws.cell(sr, 4).border = BORDER
        ws.cell(sr, 5).fill = PatternFill('solid', fgColor='E5E7EB')
        ws.cell(sr, 5).border = BORDER
        r += 1
        ws.cell(r, 1, merc).border = BORDER
        ws.cell(r, 2, 'Remanente sin asignar (stock general, para mañana)').border = BORDER
        ws.cell(r, 2).font = Font(italic=True)
        ws.cell(r, 4, remanente).border = BORDER
        ws.cell(r, 4).alignment = CENTER
        ws.cell(r, 4).font = Font(bold=True)
        r += 1
        total_cupo_gral += subtotal_cupo
        total_asig_gral += subtotal_asig
        ws.merge_cells(start_row=merc_start, start_column=1, end_row=r - 1, end_column=1)
        ws.cell(merc_start, 1).alignment = Alignment(vertical='center', horizontal='center')
        r += 1

    ws.cell(r, 2, 'TOTAL GENERAL (todas las mercaderías)').font = Font(bold=True, color='FFFFFF')
    for c in range(1, 6):
        ws.cell(r, c).fill = PatternFill('solid', fgColor='1F2937')
        ws.cell(r, c).border = BORDER
    ws.cell(r, 3, total_cupo_gral).font = Font(bold=True, color='FFFFFF')
    ws.cell(r, 3).alignment = CENTER
    ws.cell(r, 4, total_asig_gral).font = Font(bold=True, color='FFFFFF')
    ws.cell(r, 4).alignment = CENTER

    ws.column_dimensions['A'].width = 14
    ws.column_dimensions['B'].width = 32
    for col in 'CDE':
        ws.column_dimensions[col].width = 16


DIAS_PY_A_ES = ['Lunes', 'Martes', 'Miércoles', 'Jueves', 'Viernes', 'Sábado', 'Domingo']


def run(out_path=None, dia=None, wb=None, stock_extra=None, cuota_semana=None, tipif_override=None):
    """Si se pasa wb (un Workbook ya abierto, ej. desde un consolidador), agrega las hojas
    ahí en vez de crear+guardar un archivo propio — el llamador decide cuándo guardar.
    stock_extra: filas adicionales (mismo shape que read_stock_bd) para sumar al stock real
    leído de STOCK (BD) — uso: faena ya ejecutada pero todavía sin cargar ahí (ver
    stock_real.filas_manuales). cuota_semana: {'Capón': {codigo: total pedido esta semana},
    'Chancha': {...}} — ver HistoricalData.cupo_o_cuota.
    tipif_override: {correlativo: tipif_code} — parchea la TIPIFICACIÓN OP de filas que YA
    están en STOCK (BD) con Kg pero sin tipificación cargada todavía (ej. faena de hoy leída
    del romaneo en papel antes de que alguien la tipee en la planilla compartida) — no crea
    filas nuevas, solo completa el campo 'tipif' de las que ya existen por correlativo."""
    cuota_semana = cuota_semana or {}
    if dia is None:
        # hoy = día de faena; la asignación que se decide hoy es PARA el reparto del día
        # hábil siguiente (Tomás, 2026-07-31) — no para "hoy"
        fecha_reparto = dia_de_reparto(datetime.date.today())
        dia = DIAS_PY_A_ES[fecha_reparto.weekday()]
    standalone = wb is None
    hist = HistoricalData(HIST_XLSX)
    hist.redirigir('V04', 'V11')  # Tomás, 2026-08-02: V04 no trabaja más, cartera pasa a V11 "por ahora"
    stock_rows = read_stock_bd(STOCK_XLSX)
    if tipif_override:
        for r in stock_rows:
            if r['correlativo'] in tipif_override:
                r['tipif'] = tipif_override[r['correlativo']]
    if stock_extra:
        stock_rows = stock_rows + stock_extra
    if standalone:
        wb = openpyxl.Workbook()
        wb.remove(wb.active)

    shares_c, target_c, matrix_c, asignado_c, disp_c, asig_c, sobrante_c, cupo_c = asignar_capon(
        hist, stock_rows, dia=dia, cuota_semana=cuota_semana.get('Capón'))
    shares_h, target_h, matrix_h, asignado_h, disp_h, asig_h, sobrante_h, cupo_h = asignar_chancha(
        hist, stock_rows, dia=dia, cuota_semana=cuota_semana.get('Chancha'))

    ws0 = wb.create_sheet('Resumen del día')
    write_resumen_dia(ws0, dia, {
        'Capón': (shares_c, cupo_c, target_c, disp_c, asig_c),
        'Chancha': (shares_h, cupo_h, target_h, disp_h, asig_h),
    })

    ws1 = wb.create_sheet('Capón - Criterio')
    write_resumen(ws1, 'Capón', disp_c, asig_c, dia, shares_c, target_c, matrix_c, con_peso=True)
    ws2 = wb.create_sheet('Capón - Correlativos')
    write_correlativos(ws2, 'Capón', asignado_c, shares_c)
    ws2b = wb.create_sheet('Capón - Sobrante')
    write_sobrante(ws2b, 'Capón', sobrante_c)

    ws3 = wb.create_sheet('Chancha - Criterio')
    write_resumen(ws3, 'Chancha', disp_h, asig_h, dia, shares_h, target_h, matrix_h, con_peso=False)
    ws4 = wb.create_sheet('Chancha - Correlativos')
    write_correlativos(ws4, 'Chancha', asignado_h, shares_h)
    ws4b = wb.create_sheet('Chancha - Sobrante')
    write_sobrante(ws4b, 'Chancha', sobrante_h)

    hojas = {'resumen': [ws0], 'detalle': [ws2, ws4], 'sobrante': [ws2b, ws4b], 'criterio': [ws1, ws3],
             'matriz_dia': {'dia': dia, 'Capón': (shares_c, target_c), 'Chancha': (shares_h, target_h)},
             'disponible_dia': {'Capón': disp_c, 'Chancha': disp_h}, 'hist': hist}

    if standalone:
        # orden final: resumen -> correlativos (detalle por vendedor) -> sobrante -> al fondo, criterio
        def _orden(ws):
            if ws.title == 'Resumen del día':
                return 0
            if ws.title.endswith('Correlativos'):
                return 1
            if ws.title.endswith('Sobrante'):
                return 2
            return 3  # '... - Criterio'
        wb._sheets.sort(key=_orden)
        wb.save(out_path)
        print('wrote', out_path, '| dia:', dia)

    print('Capón: disponible', disp_c, '-> asignado', asig_c, '(sobran', len(sobrante_c), ')')
    print('Chancha: disponible', disp_h, '-> asignado', asig_h, '(sobran', len(sobrante_h), ')')
    return hojas


if __name__ == '__main__':
    ap = argparse.ArgumentParser()
    ap.add_argument('--out', default=None)
    ap.add_argument('--dia', default=None, choices=['Lunes', 'Martes', 'Miércoles', 'Jueves', 'Viernes'],
                     help='Default: el día de hoy (Lunes a Viernes). Forzar otro día si hace falta.')
    args = ap.parse_args()
    out = args.out or f'{OUT_DIR}\\Propuesta - Stock real - PORCINO - {datetime.date.today().isoformat()}.xlsx'
    run(out, dia=args.dia)
