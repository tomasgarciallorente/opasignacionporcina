# -*- coding: utf-8 -*-
"""Parsers para la página 'Actualizar datos' — reciben un archivo subido (file-like, de
st.file_uploader) y devuelven filas listas para guardar en Supabase (ver db.py). Reusan la
lógica de lectura del motor (stock_real.py, asignacion_engine.py) pero SIN depender de rutas
locales (OP_CARGAS de stock_real.comprometido() vive en la máquina de Tomás, no en el
servidor de Streamlit) — para el stock comprometido, en esta app se usa solo la señal de
Observaciones ('SALE ...'), no el cruce contra la lista de proveedores reales. Simplificación
deliberada para la v1: cubre el caso real encontrado hoy (Rubiolo), a costa de ser un poco
menos estricta que el motor de escritorio."""
import sys
import os
import datetime

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

import openpyxl  # noqa: E402
from stock_real import peso_bin, clean_tipif  # noqa: E402
from asignacion_engine import DIAS_SEMANA  # noqa: E402


def parse_stock_bd(file) -> list[dict]:
    """Mismo formato de hoja que STOCK (BD) de 2B-OP Stock y Entregas Porcino 2026.xlsx —
    header en la fila donde la columna D dice 'CORRELATIVO'. Devuelve filas listas para
    db.replace_stock_snapshot().

    read_only=True + iter_rows() en vez de load_workbook() normal + ws.cell() por celda
    (Tomás, 2026-08-13: "es muy lenta la carga") — el Excel de Dropbox pesa ~6MB con estilos
    de todo el libro; cargarlo entero y acceder celda a celda tardaba MÁS DE 100 SEGUNDOS
    (medido). read_only con iter_rows() salta el parseo de estilos/formato y termina en
    fracciones de segundo — mismo resultado, sin reescribir la lógica de filtrado."""
    wb = openpyxl.load_workbook(file, data_only=True, read_only=True)
    ws = wb['STOCK (BD)'] if 'STOCK (BD)' in wb.sheetnames else wb.active
    all_rows = list(ws.iter_rows(values_only=True))
    wb.close()

    def _get(row, idx):
        return row[idx] if idx < len(row) else None

    header_idx = None
    for i, row in enumerate(all_rows[:20]):
        if str(_get(row, 3) or '').strip().upper() == 'CORRELATIVO':
            header_idx = i
            break
    if header_idx is None:
        raise ValueError("No encontré la fila de header (columna D = 'CORRELATIVO'). ¿Es la hoja STOCK (BD)?")

    rows = []
    for row in all_rows[header_idx + 1:]:
        correlativo = _get(row, 3)
        kg = _get(row, 4)
        if not correlativo or not isinstance(kg, (int, float)) or kg <= 0:
            continue
        merc_raw = str(_get(row, 6) or '').strip().upper()
        if merc_raw in ('CA', 'MEI'):
            merc = 'Capón'
        elif merc_raw == 'CH':
            merc = 'Chancha'
        else:
            continue
        observaciones = _get(row, 13)
        comprometido = bool(observaciones and 'sale' in str(observaciones).lower())
        fecha = _get(row, 1)
        tipif_raw = _get(row, 2)
        rows.append({
            'correlativo': int(correlativo),
            'proveedor': str(_get(row, 0) or '').strip() or None,
            'fecha_faena': fecha.date().isoformat() if hasattr(fecha, 'date') else None,
            'kg': float(kg),
            'mercaderia': merc,
            'nivel_grasa': clean_tipif(tipif_raw) if tipif_raw is not None else None,
            'comprometido': comprometido,
        })
    return rows


def _parse_fecha_pegada(v):
    """La fecha llega como texto (pegado desde Excel en una celda de tabla editable de
    Streamlit) o como datetime si alguna vez se pasa un valor ya parseado."""
    if v is None or v == '':
        return None
    if hasattr(v, 'date'):
        return v.date().isoformat()
    texto = str(v).strip()
    for fmt in ('%d/%m/%Y', '%d/%m/%y', '%Y-%m-%d'):
        try:
            return datetime.datetime.strptime(texto, fmt).date().isoformat()
        except ValueError:
            continue
    return None


def _fila_desde_valores(cols_stock, valores):
    """valores: lista posicional (mismo orden que cols_stock) -> dict con los mismos nombres
    de columna que usa un DataFrame bien pegado, para reusar la misma extracción de abajo."""
    return {cols_stock[i]: (valores[i] if i < len(valores) else None) for i in range(len(cols_stock))}


def _reconstruir_filas_pegado_roto(df, cols_stock):
    """A veces el data_editor de Streamlit no separa el pegado en celdas — todo el bloque
    copiado del Excel (varias filas y columnas, tabuladas) cae entero como texto suelto en
    UNA sola celda (Tomás, 2026-08-13: 'copio y pego del excel y lo pega así y queda mal').
    Detecta esas celdas (contienen tabs o saltos de línea) y las vuelve a partir a mano por
    línea y por tab, en el mismo orden de columnas — así el resultado es idéntico a si el
    pegado se hubiera separado bien de entrada."""
    filas = []
    for _, fila in df.iterrows():
        bloque = None
        for v in fila.values:
            if isinstance(v, str) and ('\t' in v or '\n' in v):
                bloque = v
                break
        if bloque is None:
            continue
        for linea in bloque.splitlines():
            if not linea.strip():
                continue
            filas.append(_fila_desde_valores(cols_stock, linea.split('\t')))
    return filas


def _extraer_fila_stock(fila):
    """Toma un dict/Serie con las columnas del pegado y devuelve la fila lista para
    db.replace_stock_snapshot(), o None si no tiene Correlativo/Kg válidos."""
    correlativo = fila.get('Correlativo')
    kg = fila.get('Kg')
    if correlativo in (None, '') or kg in (None, ''):
        return None
    try:
        correlativo = int(float(str(correlativo).strip()))
        kg = float(str(kg).strip())
    except (TypeError, ValueError):
        return None
    if kg <= 0:
        return None
    merc_raw = str(fila.get('Mercadería') or '').strip().upper()
    if merc_raw in ('CA', 'MEI'):
        merc = 'Capón'
    elif merc_raw == 'CH':
        merc = 'Chancha'
    else:
        return None
    observaciones = fila.get('Observaciones')
    comprometido = bool(observaciones and 'sale' in str(observaciones).lower())
    tipif_raw = fila.get('Tipificación OP')
    return {
        'correlativo': correlativo,
        'proveedor': str(fila.get('Productor') or '').strip() or None,
        'fecha_faena': _parse_fecha_pegada(fila.get('Fecha faena')),
        'kg': kg,
        'mercaderia': merc,
        'nivel_grasa': clean_tipif(tipif_raw) if tipif_raw not in (None, '') else None,
        'comprometido': comprometido,
    }


def parse_stock_pegado(df) -> list[dict]:
    """df: pandas.DataFrame pegado a mano en la grilla de 'Actualizar datos' — Tomás,
    2026-08-13: 'quiero poder copiar y pegar desde el archivo Excel, no quiero subir el
    archivo'. Mismas columnas y mismo mapeo/filtrado que parse_stock_bd() (Productor, Fecha
    faena, Tipificación OP, Correlativo, Kg, X, Mercadería, Conf., Gras., Garrón, Tropa, Cat,
    Calidad, Observaciones) para que un pegado de A:N completo entre tal cual.

    Primero intenta leer el DataFrame tal cual (pegado que sí se separó bien en celdas); lo
    que no haya dado ningún animal válido ahí, lo reintenta reconstruyendo filas rotas (ver
    _reconstruir_filas_pegado_roto) — cubre los dos casos sin que el usuario tenga que saber
    cuál pasó."""
    rows = []
    correlativos_vistos = set()
    for _, fila in df.iterrows():
        r = _extraer_fila_stock(fila)
        if r and r['correlativo'] not in correlativos_vistos:
            rows.append(r)
            correlativos_vistos.add(r['correlativo'])

    cols_stock = list(df.columns)
    for fila in _reconstruir_filas_pegado_roto(df, cols_stock):
        r = _extraer_fila_stock(fila)
        if r and r['correlativo'] not in correlativos_vistos:
            rows.append(r)
            correlativos_vistos.add(r['correlativo'])

    return rows


def parse_historico(file) -> tuple[list[dict], list[dict]]:
    """Mismo formato que 'Promedio diario Capón y Chancha por vendedor.xlsx' — reusa
    HistoricalData (asignacion_engine.py) para no reimplementar el parseo, y aplana el
    resultado a filas para db.replace_historico()."""
    from asignacion_engine import HistoricalData
    hist = HistoricalData(xlsx_path=file)

    bloque_rows = []
    for merc, plan in [('Capón', hist.plan_capon), ('Chancha', hist.plan_chancha)]:
        for code, name, _row, dias in plan:
            for dia in DIAS_SEMANA + ['Total']:
                bloque_rows.append({
                    'bloque_codigo': code, 'bloque_nombre': name, 'mercaderia': merc,
                    'dia': dia, 'cupo': float(dias.get(dia, 0.0)),
                })

    calidad_rows = []
    for merc, eje, pct_dict in [
        ('Capón', 'peso', hist.peso_pct),
        ('Capón', 'tipif', hist.tipif_pct_capon),
        ('Chancha', 'tipif', hist.tipif_pct_chancha),
    ]:
        for code, categorias in pct_dict.items():
            for categoria, pct in categorias.items():
                calidad_rows.append({
                    'bloque_codigo': code, 'mercaderia': merc, 'eje': eje,
                    'categoria': categoria, 'pct': float(pct),
                })
    return bloque_rows, calidad_rows
