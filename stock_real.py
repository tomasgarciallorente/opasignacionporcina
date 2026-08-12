# -*- coding: utf-8 -*-
"""Lee el stock disponible REAL, carcaza por carcaza, de la hoja STOCK (BD) de
2B-OP Stock y Entregas Porcino 2026.xlsx (Productor, Fecha faena, Tipificación OP,
Correlativo, Kg, Mercadería — fila 8 header, datos desde fila 9)."""
import re
import openpyxl

OP_CARGAS = (r'C:\Users\Gtecomercial\Dropbox\01-Abasto\01-Cargas de hacienda'
             r'\00-Planilla de cargas B+P\OP Cargas bovinas y porcinas 2026.xlsx')


def _normalizar_proveedor(s):
    """minúsculas, sin espacios/puntuación — para cruzar 'Bassano SA' con 'Bassano S.A.'."""
    return re.sub(r'[^a-z0-9]', '', str(s or '').lower())


def proveedores_porcino_conocidos(path=OP_CARGAS):
    """Lista real de proveedores de hacienda porcina (hoja 'Proveed. P' de OP Cargas) —
    Tomás, 2026-08-11: "Rubiolo es cliente, no es proveedor". En STOCK (BD) la columna
    PRODUCTOR a veces trae el nombre de un CLIENTE en vez de un proveedor real, cuando esa
    mercadería ya está comprometida para salir a ese cliente puntual (ver comprometido() más
    abajo) — no es stock libre para prorratear entre bloques."""
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb['Proveed. P']
    out = set()
    for row in ws.iter_rows(min_row=3, max_col=1, values_only=True):
        v = row[0]
        if v and str(v).strip() not in ('.', ''):
            out.add(_normalizar_proveedor(v))
    return out


def comprometido(observaciones, productor, proveedores_conocidos):
    """True si la fila de STOCK (BD) ya está comprometida a un cliente puntual, no es stock
    libre para el reparto por bloques — Tomás, 2026-08-11, caso Rubiolo (48 carcazas con
    Observaciones='SALE MIERCOLES' y Productor='RUBIOLO', un cliente, no un proveedor real).
    Dos señales combinadas (ambas se chequean, cualquiera de las dos excluye la fila):
    (1) Observaciones menciona una salida programada ('SALE ...'); (2) Productor no figura en
    la lista real de proveedores de OP Cargas — puede ser un cliente, o simplemente ruido."""
    if observaciones and 'sale' in str(observaciones).lower():
        return True
    if _normalizar_proveedor(productor) not in proveedores_conocidos:
        return True
    return False

WEIGHT_BINS = [
    ('< 71 kg', lambda kg: kg < 71),
    ('71 a 95 kg', lambda kg: 71 <= kg <= 95),
    ('96 a 105 kg', lambda kg: 96 <= kg <= 105),
    ('106 a 115 kg', lambda kg: 106 <= kg <= 115),
    ('116 a 125 kg', lambda kg: 116 <= kg <= 125),
    ('>= 126 kg', lambda kg: kg >= 126),
]
TIPIF_CATS = ['-1', '0', '1', '1+', '1++', '2', '2+', '3', '3+', '4']
TIPIF_OTHER = 'Otro / sin clasif.'
TIPIF_TOKEN_RE = re.compile(r'(?<![0-9A-Za-z])(-?\d(?:\+\+|\+|-)?)(?![0-9A-Za-z])')


def peso_bin(kg):
    for label, pred in WEIGHT_BINS:
        if pred(kg):
            return label
    return None


def clean_tipif(raw):
    s = str(raw).strip() if raw is not None else ''
    if not s:
        return TIPIF_OTHER
    m = TIPIF_TOKEN_RE.search(s)
    if not m:
        return TIPIF_OTHER
    code = m.group(1)
    return code if code in TIPIF_CATS else TIPIF_OTHER


def read_stock_bd(path, excluir_comprometidos=True):
    """Devuelve lista de dicts: correlativo, proveedor, fecha_faena, tipif, kg, peso, merc
    (merc = 'Capón' para CA/MEI, 'Chancha' para CH; el resto —LECHON/PA— se ignora).
    excluir_comprometidos: saca del pool las filas ya comprometidas a un cliente puntual (ver
    comprometido()) — activado por default, es lo correcto para el reparto por bloques."""
    proveedores_conocidos = proveedores_porcino_conocidos() if excluir_comprometidos else None
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb['STOCK (BD)']
    # localizar la fila de header (columna D = 'CORRELATIVO') en vez de hardcodear el número de
    # fila, por si la tabla síntesis de arriba cambia de tamaño en el futuro.
    header_row = None
    for r in range(1, 20):
        if str(ws.cell(r, 4).value or '').strip().upper() == 'CORRELATIVO':
            header_row = r
            break
    if header_row is None:
        raise SystemExit('No encontré la fila de header (columna D = CORRELATIVO) en STOCK (BD).')

    # OJO (2026-08-06): la tabla NO es un bloque contiguo — es una lista de proveedores con
    # filas placeholder en blanco para los que no tienen stock en ese momento, intercaladas con
    # filas de datos reales para los que sí. Cortar en el primer hueco en blanco (como se hacía
    # antes) se comía datos reales que aparecen MÁS ABAJO, después del hueco — bug confirmado:
    # se perdían 2 chanchas de "Don Ramón" (05/08) porque quedaban debajo de un hueco de
    # proveedores sin stock. Ahora se recorre TODA la tabla hasta el final real de la hoja
    # (ws.max_row) y cada fila en blanco simplemente se salta, no corta la lectura.
    rows = []
    for r in range(header_row + 1, ws.max_row + 1):
        correlativo = ws.cell(r, 4).value
        kg = ws.cell(r, 5).value
        if not correlativo or not isinstance(kg, (int, float)) or kg <= 0:
            continue
        productor = ws.cell(r, 1).value
        observaciones = ws.cell(r, 14).value
        if excluir_comprometidos and comprometido(observaciones, productor, proveedores_conocidos):
            continue
        merc_raw = str(ws.cell(r, 7).value or '').strip().upper()
        if merc_raw in ('CA', 'MEI'):
            merc = 'Capón'
        elif merc_raw == 'CH':
            merc = 'Chancha'
        else:
            continue  # LECHON, PA u otro — fuera de alcance hoy (conteo 0 al 2026-07-29)
        fecha = ws.cell(r, 2).value
        rows.append({
            'correlativo': correlativo,
            'proveedor': productor,
            'fecha_faena': fecha,
            'tipif_raw': ws.cell(r, 3).value,
            'tipif': clean_tipif(ws.cell(r, 3).value),
            'kg': float(kg),
            'peso': peso_bin(float(kg)),
            'merc': merc,
            'es_real': True,
        })
    return rows


def filas_manuales(entries, prefijo='PEND', es_real=True):
    """Filas sintéticas sin correlativo/peso/tipif real: dos usos distintos con el mismo shape.
    (1) prefijo='PEND', es_real=True — faena YA ejecutada pero todavía sin tipificación cargada
    en STOCK (BD) (Tomás, 2026-08-02: "la faena de ayer, por ahora no tenemos la tipificación").
    (2) prefijo='PROY', es_real=False — faena que TODAVÍA no se hizo, viene del plan de Compras
    (Tomás, 2026-08-04: el stock real hay que consumirlo entero día a día hasta agotarlo, recién
    ahí sigue lo proyectado — estas filas entran al MISMO pool que las reales, más nuevas en
    fecha, así el FIFO las deja para después de lo que ya existe físicamente).
    entries: [{'proveedor':.., 'fecha_faena': date, 'merc': 'Capón'/'Chancha', 'cantidad': int}].
    Cada cabeza queda con tipif='Otro / sin clasif.' y peso=None — la Capa 2/3 las reparte por
    el peso relativo general del bloque (red de seguridad ya existente en
    reconciliar_con_stock_real) hasta que haya dato real.
    Correlativo 'PROY-{merc}-{n}' con numeración corrida (no repite por fecha) — la fecha de
    faena real va en su propia columna en la hoja de detalle, no hace falta repetirla adentro
    del nombre del correlativo (Tomás, 2026-08-06: ej. 'PROY-Capón-47')."""
    rows = []
    n = 0
    for e in entries:
        for _i in range(int(e['cantidad'])):
            n += 1
            rows.append({
                'correlativo': f"{prefijo}-{e['merc']}-{n}",
                'proveedor': e['proveedor'],
                'fecha_faena': e['fecha_faena'],
                'tipif_raw': None,
                'tipif': TIPIF_OTHER,
                'kg': None,
                'peso': None,
                'merc': e['merc'],
                'es_real': es_real,
            })
    return rows
