# -*- coding: utf-8 -*-
"""Herramienta puntual de asignación de mercadería porcina — Origen Pampa.
Login simple (contraseña compartida) -> Cargar tipificación -> Generar reparto ->
Actualizar datos. Ver plan: C:\\Users\\Gtecomercial\\.claude\\plans\\moonlit-prancing-willow.md
"""
import datetime
import hashlib
import io
import os

import pandas as pd
import streamlit as st

import db
import excel_import
from motor_adapter import generar_reparto, generar_reparto_semanal, construir_stock_rows
from run_asignacion_stock_real import write_resumen_dia, write_correlativos, write_sobrante
from asignacion_engine import dia_de_reparto
import openpyxl

LOGO_PATH = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'logo.png')

st.set_page_config(page_title='Asignación porcina — Origen Pampa', page_icon=LOGO_PATH, layout='wide')

OBSERVACIONES = [
    'Ninguna', 'Golpe en cadera', 'Golpe interno', 'Golpe interno feo',
    'Corte en paleta (cuero)', 'Otro',
]
GRADOS = ['-1', '0', '1', '1+', '1++', '2', '2+', '3', '3+', '4']
DIAS_HABILES = ['Lunes', 'Martes', 'Miércoles', 'Jueves', 'Viernes']
DIAS_PY_A_ES = ['Lunes', 'Martes', 'Miércoles', 'Jueves', 'Viernes', 'Sábado', 'Domingo']

CSS = """
<style>
[data-testid="stMetricValue"] { color: #2b1414; }
[data-testid="stMetric"] {
    background: #ffffff; border: 1px solid rgba(204,26,26,0.25); border-radius: 8px; padding: 10px;
}
h1, h2, h3 { color: #2b1414 !important; }
.op-header { display: flex; align-items: center; gap: 14px; margin-bottom: 4px; }
.op-header img { width: 44px; height: 44px; border-radius: 50%; }
.op-header .op-title { font-size: 1.4rem; font-weight: 700; color: #2b1414; }
.op-header .op-sub { color: #7a4a4a; font-size: 0.85rem; }
.op-pill {
    display: inline-block; background: #cc1a1a; color: #f5efef; border-radius: 999px;
    padding: 2px 12px; font-size: 0.8rem; font-weight: 600; margin-left: 8px;
}
</style>
"""


@st.fragment(run_every=60)
def render_header():
    """Fragment aparte (Tomás, 2026-08-12: "actualizá hora también" / "quiero que
    constantemente se actualice") — se re-renderiza solo, cada 60s, sin recargar el resto de
    la página (no pisa formularios abiertos en otras secciones)."""
    st.markdown(CSS, unsafe_allow_html=True)
    ahora = datetime.datetime.now()
    hoy = ahora.date()
    semana_iso = hoy.isocalendar()[1]
    dia_semana = DIAS_PY_A_ES[hoy.weekday()] if hoy.weekday() < 7 else ''
    col_logo, col_txt = st.columns([1, 8])
    with col_logo:
        st.image(LOGO_PATH, width=56)
    with col_txt:
        st.markdown(
            f'<div class="op-title">Origen Pampa — Asignación porcina'
            f'<span class="op-pill">Semana {semana_iso}</span></div>'
            f'<div class="op-sub">Hoy {dia_semana} {hoy.strftime("%d/%m/%Y")} — {ahora.strftime("%H:%M")}hs</div>',
            unsafe_allow_html=True,
        )
    st.divider()


# --- Login ---
def _token_login():
    """Token derivado de la contraseña (no la contraseña en texto plano) para guardar el
    login en la URL — Tomás, 2026-08-13: 'cada vez que le doy refresh debo poner de nuevo la
    contraseña'. Streamlit arranca sesión nueva en cada F5 (pierde session_state), pero los
    query params sobreviven al refresh porque son parte de la URL."""
    pw = st.secrets.get('APP_PASSWORD')
    return hashlib.sha256(pw.encode()).hexdigest()[:20] if pw else None


def check_password():
    if st.session_state.get('autenticado'):
        return True
    token_esperado = _token_login()
    if token_esperado and st.query_params.get('k') == token_esperado:
        st.session_state['autenticado'] = True
        return True
    st.markdown(CSS, unsafe_allow_html=True)
    col1, col2 = st.columns([1, 3])
    with col1:
        st.image(LOGO_PATH, width=90)
    with col2:
        st.markdown('<div class="op-title" style="font-size:1.8rem;">Origen Pampa</div>'
                     '<div class="op-sub">Asignación porcina</div>', unsafe_allow_html=True)
    pw = st.text_input('Contraseña', type='password')
    if st.button('Entrar'):
        if pw == st.secrets.get('APP_PASSWORD'):
            st.session_state['autenticado'] = True
            st.query_params['k'] = token_esperado
            st.rerun()
        else:
            st.error('Contraseña incorrecta.')
    return False


if not check_password():
    st.stop()

render_header()

# --- Navegación ---
pagina = st.sidebar.radio('Página', ['Cargar tipificación', 'Generar reparto', 'Actualizar datos'])
def _cerrar_sesion():
    st.session_state.pop('autenticado', None)
    st.query_params.pop('k', None)


st.sidebar.button('Cerrar sesión', on_click=_cerrar_sesion)


# ============================== CARGAR TIPIFICACIÓN ==============================
if pagina == 'Cargar tipificación':
    st.title('Cargar tipificación de una tropa')

    if 'tropa_generada' not in st.session_state:
        st.session_state['tropa_generada'] = None

    with st.form('form_tropa'):
        col1, col2 = st.columns(2)
        proveedor = col1.text_input('Proveedor', placeholder='Ej. ISOWEAN S.A.')
        fecha_faena = col2.date_input('Fecha de faena', value=datetime.date.today())
        col3, col4, col5 = st.columns(3)
        mercaderia = col3.radio('Mercadería', ['Capón', 'Chancha'], horizontal=True)
        correlativo_inicial = col4.number_input('Correlativo inicial', min_value=1, step=1)
        cantidad = col5.number_input('Cantidad de cabezas', min_value=1, max_value=500, step=1)
        generar = st.form_submit_button('Generar planilla de carga')

    if generar:
        st.session_state['tropa_generada'] = {
            'proveedor': proveedor, 'fecha_faena': fecha_faena, 'mercaderia': mercaderia,
            'correlativo_inicial': int(correlativo_inicial), 'cantidad': int(cantidad),
        }

    tropa = st.session_state['tropa_generada']
    if tropa:
        st.divider()
        st.subheader(f"{tropa['mercaderia']} — {tropa['cantidad']} cabezas, correlativo {tropa['correlativo_inicial']}..{tropa['correlativo_inicial'] + tropa['cantidad'] - 1}")
        st.caption('Tabla tipo Excel — tocá una celda para editarla, o pegá una columna copiada de '
                   'otra planilla (Ctrl+V) para cargar varias filas de una. Grado y Observación se '
                   'eligen de una lista desplegable.')

        # Mismo orden de columnas que el romaneo en papel / STOCK (BD) del Excel de Dropbox
        # (Correlativo, Garrón, Kg, Tipificación, Observaciones) — Tomás, 2026-08-13: "que
        # tenga el mismo formato... así es más fácil copiar y pegar desde el archivo". Garrón
        # es solo de referencia visual para alinear el pegado, no se guarda (no forma parte
        # del motor de reparto).
        df_tropa = pd.DataFrame({
            'Correlativo': [tropa['correlativo_inicial'] + i for i in range(tropa['cantidad'])],
            'Garrón': [0] * tropa['cantidad'],
            'Kg': [0.0] * tropa['cantidad'],
            'Grado': ['1'] * tropa['cantidad'],
            'Observación': ['Ninguna'] * tropa['cantidad'],
        })
        df_tropa_editado = st.data_editor(
            df_tropa, key='editor_tipificacion', use_container_width=True, hide_index=True,
            height=min(35 * (tropa['cantidad'] + 1) + 3, 900),
            column_config={
                'Correlativo': st.column_config.NumberColumn('Correlativo', disabled=True),
                'Garrón': st.column_config.NumberColumn('Garrón', min_value=0, step=1, help='Solo de referencia, no se guarda.'),
                'Kg': st.column_config.NumberColumn('Kg', min_value=1.0, max_value=300.0, step=1.0),
                'Grado': st.column_config.SelectboxColumn('Grado', options=GRADOS),
                'Observación': st.column_config.SelectboxColumn('Observación', options=OBSERVACIONES),
            },
        )
        filas = [{
            'correlativo': int(fila['Correlativo']), 'kg': float(fila['Kg']), 'nivel_grasa': fila['Grado'],
            'observacion': None if fila['Observación'] == 'Ninguna' else fila['Observación'],
        } for _, fila in df_tropa_editado.iterrows()]

        st.divider()
        fotos = st.file_uploader('Fotos del romaneo (opcional, respaldo de toda la tropa — podés subir varias)',
                                  type=['png', 'jpg', 'jpeg'], accept_multiple_files=True)
        creado_por = st.text_input('Tu nombre', placeholder='Ej. Ignacio Beas')

        if st.button('Confirmar y guardar', type='primary'):
            sin_kg = [f['correlativo'] for f in filas if f['kg'] <= 0]
            if not tropa['proveedor']:
                st.error('Falta el proveedor.')
            elif not creado_por:
                st.error('Falta tu nombre.')
            elif sin_kg:
                st.error(f"Faltan cargar los Kg de {len(sin_kg)} animal(es): {', '.join(map(str, sin_kg[:15]))}"
                          + ('...' if len(sin_kg) > 15 else ''))
            else:
                lote_id = db.insert_lote(tropa['proveedor'], tropa['fecha_faena'], tropa['mercaderia'],
                                          tropa['correlativo_inicial'], tropa['cantidad'], None, creado_por)
                db.insert_animales(lote_id, filas)
                if fotos:
                    urls = [db.upload_foto(f.getvalue(), f.name, f.type) for f in fotos]
                    db.insert_fotos(lote_id, urls)
                st.success(f"Guardado — {tropa['cantidad']} animales cargados"
                           + (f", {len(fotos)} foto(s)." if fotos else "."))
                st.session_state['tropa_generada'] = None
                st.rerun()


# ============================== GENERAR REPARTO ==============================
elif pagina == 'Generar reparto':
    st.title('Generar reparto por bloques')

    bloque_rows, calidad_rows = db.fetch_historico()
    if not bloque_rows:
        st.warning('Todavía no se subió el Excel histórico (Promedio diario Capón y Chancha por '
                   'vendedor.xlsx) — andá a "Actualizar datos" antes de generar un reparto.')
        st.stop()

    # --- Estado del stock, deduplicado de verdad (Tomás, 2026-08-12: "está poniendo un stock
    # que no existe" — el total anterior sumaba snapshot + tipificación sin descontar los
    # correlativos que aparecen en ambos lados) ---
    @st.fragment(run_every=60)
    def render_stock_panel():
        """Fragment aparte (Tomás, 2026-08-12: "quiero que constantemente se actualice el
        stock") — vuelve a consultar Supabase y se re-renderiza solo cada 60s, sin recargar
        el resto de la página. Guarda el resultado en session_state para que el botón
        'Generar propuesta' (fuera del fragment) siempre lea el último dato consultado."""
        snap_info = db.fetch_stock_snapshot_info()
        lotes_hoy = db.fetch_tipificacion_hoy_info()
        snapshot_rows = db.fetch_stock_snapshot()
        # Tomás, 2026-08-13: "al pegarte esa tabla de stock considerá que es el total
        # disponible. Por ahora esa va a ser la manera de trabajar, después vamos a explorar
        # cargando la tipificación" — por ahora NO se suma la tipificación cargada aparte
        # (evita el doble conteo/confusión de antes); el snapshot pegado en "Actualizar datos"
        # es la única fuente de stock. Se deja construir_stock_rows tal cual (acepta lista
        # vacía) para no tener que tocar su lógica cuando se retome la carga de tipificación.
        animales_tipificados = []
        ya_repartidos = db.fetch_correlativos_ya_repartidos()
        stock_rows = construir_stock_rows(snapshot_rows, animales_tipificados, ya_repartidos=ya_repartidos)
        st.session_state['stock_actual'] = {
            'snapshot_rows': snapshot_rows, 'animales_tipificados': animales_tipificados,
            'ya_repartidos': ya_repartidos, 'stock_rows': stock_rows,
        }
        animales_hoy = sum(l['cantidad'] for l in lotes_hoy)
        n_capon = sum(1 for r in stock_rows if r['merc'] == 'Capón')
        n_chancha = sum(1 for r in stock_rows if r['merc'] == 'Chancha')

        with st.container(border=True):
            col_titulo, col_boton = st.columns([5, 1])
            with col_titulo:
                st.markdown('**Stock disponible ahora mismo (deduplicado, sin lo ya repartido y guardado)** '
                             '<span style="color:#7a4a4a; font-size:0.8rem;">— se actualiza solo cada 1 '
                             f'min, última consulta {datetime.datetime.now().strftime("%H:%M:%S")}</span>',
                             unsafe_allow_html=True)
            with col_boton:
                # Botón dentro del fragment: al tocarlo, Streamlit re-ejecuta SOLO este
                # fragment (no toda la página) — refresco inmediato sin esperar el timer de
                # 60s. Tomás, 2026-08-13: "quiero que se actualice el stock inmediatamente
                # así puedo chequear que esté bien".
                st.button('🔄 Actualizar', key='refresh_stock_manual', use_container_width=True)
            c1, c2, c3 = st.columns(3)
            if snap_info['ultimo_sync']:
                ts = datetime.datetime.fromisoformat(snap_info['ultimo_sync'].replace('Z', '+00:00'))
                ts_local = ts.astimezone().strftime('%d/%m %H:%M')
                c1.metric('Última sync del Excel', ts_local, f"{snap_info['total_filas']} piezas en el snapshot")
            else:
                c1.metric('Última sync del Excel', 'nunca')
            c2.metric('Tipificado hoy', f'{animales_hoy} animales', f'{len(lotes_hoy)} tropa(s)')
            c3.metric('Total disponible real', f'{len(stock_rows)}', f'{n_capon} Capón / {n_chancha} Chancha')
            # Tomás, 2026-08-13: el snapshot pegado/subido ES el total disponible — ya no se
            # suma la tipificación cargada aparte (ver comentario arriba).
            st.caption(
                f"Cómo se arma el total: {snap_info['total_filas']} del snapshot de Excel/pegado "
                f"− {len(ya_repartidos)} ya repartidos = **{len(stock_rows)} piezas disponibles reales**."
            )
            if lotes_hoy:
                st.caption('Tropas tipificadas hoy (informativo — todavía **no** se suman al '
                           'total disponible, eso lo dejamos para más adelante): ' + ', '.join(
                    f"{l['proveedor']} ({l['mercaderia']}, {l['cantidad']})" for l in lotes_hoy))
            if ya_repartidos:
                st.caption(f'{len(ya_repartidos)} piezas ya repartidas y guardadas en corridas anteriores — excluidas de este pool.')

            with st.expander(f'Ver detalle — {len(stock_rows)} piezas'):
                st.dataframe(
                    [{'Correlativo': r['correlativo'], 'Mercadería': r['merc'], 'Kg': r['kg'],
                      'Grado': r['tipif'], 'Proveedor': r['proveedor'],
                      'Fecha faena': r['fecha_faena'], 'Observación': r.get('observacion') or ''}
                     for r in sorted(stock_rows, key=lambda r: (r['merc'], r['correlativo']))],
                    use_container_width=True, hide_index=True, height=300,
                )

    render_stock_panel()
    _stock_actual = st.session_state.get('stock_actual', {})
    snapshot_rows = _stock_actual.get('snapshot_rows', [])
    animales_tipificados = _stock_actual.get('animales_tipificados', [])
    ya_repartidos = _stock_actual.get('ya_repartidos', set())
    stock_rows = _stock_actual.get('stock_rows', [])

    tab_dia, tab_semana = st.tabs(['📅 Reparto del próximo día', '🗓️ Reparto semanal'])

    # ---------------------------- REPARTO DE UN DÍA ----------------------------
    with tab_dia:
        dia_sugerido = DIAS_PY_A_ES[dia_de_reparto(datetime.date.today()).weekday()]
        idx_default = DIAS_HABILES.index(dia_sugerido) if dia_sugerido in DIAS_HABILES else 0
        dia = st.selectbox('Día de reparto', DIAS_HABILES, index=idx_default,
                            help='Por defecto, el día hábil siguiente a hoy (hoy es día de faena, '
                                 'se reparte al día siguiente).')
        st.caption('Este día respeta el cupo histórico de ESE día por bloque — no vacía todo el '
                   'stock de una. Lo que no entra en el cupo de hoy queda disponible (no se '
                   'pierde) para que una corrida de otro día lo tome, hasta agotar el 100%.')

        if st.button('Generar propuesta', type='primary'):
            if not stock_rows:
                st.error('No hay stock cargado — ni tipificación de hoy ni snapshot de Excel subido.')
                st.stop()

            resultado = generar_reparto(snapshot_rows, animales_tipificados, bloque_rows, calidad_rows, dia=dia,
                                         ya_repartidos=ya_repartidos)
            st.session_state['resultado'] = resultado

        resultado = st.session_state.get('resultado')
        if resultado:
            st.divider()
            st.caption(f"Reparto para el {resultado['dia']}")
            for merc_key, merc_label in [('capon', 'Capón'), ('chancha', 'Chancha')]:
                d = resultado[merc_key]
                st.subheader(merc_label)
                c1, c2, c3 = st.columns(3)
                c1.metric('Disponible', d['disponible'])
                c2.metric('Asignado', d['asignado_total'])
                c3.metric('Sobrante', len(d['sobrante']))

                filas_resumen = []
                for code, (nombre, _share) in d['shares'].items():
                    filas_resumen.append({'Bloque': nombre, 'Asignado': d['target'].get(code, 0)})
                st.dataframe(filas_resumen, use_container_width=True, hide_index=True)

                if d['movimientos_golpes_cortes']:
                    st.markdown('**Golpes / cortes — reasignados por afinidad a magros**')
                    st.dataframe(d['movimientos_golpes_cortes'], use_container_width=True, hide_index=True)

                with st.expander(f'Detalle de correlativos por bloque — {merc_label}'):
                    for code, (nombre, _share) in d['shares'].items():
                        animales = d['asignado'].get(code, [])
                        if not animales:
                            continue
                        st.markdown(f'**{nombre}** ({len(animales)})')
                        st.dataframe(
                            [{'Correlativo': a['correlativo'], 'Kg': a['kg'], 'Grado': a['tipif'],
                              'Observación': a.get('observacion') or ''} for a in animales],
                            use_container_width=True, hide_index=True,
                        )

            # --- Excel de descarga, mismo formato que el motor de escritorio ---
            wb = openpyxl.Workbook()
            wb.remove(wb.active)
            ws0 = wb.create_sheet('Resumen del día')
            write_resumen_dia(ws0, resultado['dia'], {
                'Capón': (resultado['capon']['shares'], resultado['capon']['cupo'], resultado['capon']['target'],
                          resultado['capon']['disponible'], resultado['capon']['asignado_total']),
                'Chancha': (resultado['chancha']['shares'], resultado['chancha']['cupo'], resultado['chancha']['target'],
                            resultado['chancha']['disponible'], resultado['chancha']['asignado_total']),
            })
            for merc_key, merc_label, con_peso in [('capon', 'Capón', True), ('chancha', 'Chancha', False)]:
                d = resultado[merc_key]
                ws_det = wb.create_sheet(f'{merc_label} - Correlativos')
                write_correlativos(ws_det, merc_label, d['asignado'], d['shares'])
                ws_sob = wb.create_sheet(f'{merc_label} - Sobrante')
                write_sobrante(ws_sob, merc_label, d['sobrante'])

            buf = io.BytesIO()
            wb.save(buf)
            st.download_button('Descargar Excel', buf.getvalue(),
                                file_name=f"Propuesta reparto - {resultado['dia']} - {datetime.date.today().isoformat()}.xlsx")

            if st.button('Guardar este resultado'):
                filas_guardar = []
                for merc_key, merc_label in [('capon', 'Capón'), ('chancha', 'Chancha')]:
                    d = resultado[merc_key]
                    movidos = {m['correlativo'] for m in d['movimientos_golpes_cortes'] if m['destino']}
                    for code, (_nombre, _s) in d['shares'].items():
                        for a in d['asignado'].get(code, []):
                            filas_guardar.append({
                                'mercaderia': merc_label, 'bloque_codigo': code, 'correlativo': a['correlativo'],
                                'kg': a['kg'], 'nivel_grasa': a['tipif'], 'observacion': a.get('observacion'),
                                'reasignado_por_golpe_corte': a['correlativo'] in movidos,
                            })
                db.insert_reparto_resultados(resultado['dia'], filas_guardar)
                st.success('Guardado.')

    # ---------------------------- REPARTO SEMANAL ----------------------------
    with tab_semana:
        st.caption('Reparte el stock disponible entre todos los días que quedan de esta semana '
                   '(desde el próximo reparto hasta el viernes), respetando el cupo de cada día — '
                   'el sobrante de un día pasa como pool al siguiente hasta agotar el 100%.')

        if st.button('Generar propuesta semanal', type='primary'):
            if not stock_rows:
                st.error('No hay stock cargado — ni tipificación de hoy ni snapshot de Excel subido.')
                st.stop()
            st.session_state['resultado_semana'] = generar_reparto_semanal(
                snapshot_rows, animales_tipificados, bloque_rows, calidad_rows, ya_repartidos=ya_repartidos)

        rs = st.session_state.get('resultado_semana')
        if rs:
            st.divider()
            st.caption(f"Semana {rs['semana_iso']} — días: {', '.join(rs['dias'])}")
            for merc_key, merc_label in [('capon', 'Capón'), ('chancha', 'Chancha')]:
                m = rs[merc_key]
                st.subheader(merc_label)

                # Tabla bloque x día
                bloques_todos = {}
                for r in m['resultados']:
                    for code, (nombre, _s) in r['shares'].items():
                        bloques_todos.setdefault(code, nombre)
                filas = []
                for code, nombre in bloques_todos.items():
                    fila = {'Bloque': nombre}
                    total = 0
                    for r in m['resultados']:
                        v = r['asignado'].get(code, 0)
                        fila[r['dia']] = v
                        total += v
                    fila['Total semana'] = total
                    filas.append(fila)
                st.dataframe(filas, use_container_width=True, hide_index=True)
                st.caption(f"Sobrante sin asignar al final de la semana: {len(m['sobrante_fin_semana'])}")

                for r in m['resultados']:
                    movimientos = m['movimientos_golpes_cortes_por_dia'].get(r['dia'], [])
                    if movimientos:
                        st.markdown(f"**Golpes/cortes — {r['dia']}**")
                        st.dataframe(movimientos, use_container_width=True, hide_index=True)

                with st.expander(f'Detalle de correlativos por día y bloque — {merc_label}'):
                    for r in m['resultados']:
                        st.markdown(f"**{r['dia']}**")
                        for code, (nombre, _s) in r['shares'].items():
                            animales = r['correlativos'].get(code, [])
                            if not animales:
                                continue
                            st.markdown(f'{nombre} ({len(animales)})')
                            st.dataframe(
                                [{'Correlativo': a['correlativo'], 'Kg': a['kg'], 'Grado': a['tipif'],
                                  'Observación': a.get('observacion') or ''} for a in animales],
                                use_container_width=True, hide_index=True,
                            )

            # --- Excel de descarga, un libro con el resumen semanal + un juego de solapas
            # por día (mismo formato que el reparto diario, reusando write_resumen_dia /
            # write_correlativos / write_sobrante) ---
            wb = openpyxl.Workbook()
            ws_resumen = wb.active
            ws_resumen.title = 'Resumen semana'
            ws_resumen['A1'] = f"Resumen semana {rs['semana_iso']} — días: {', '.join(rs['dias'])}"
            # Tabla simple, una sección por mercadería, una fila por bloque.
            row_cursor = 3
            for merc_key, merc_label in [('capon', 'Capón'), ('chancha', 'Chancha')]:
                m = rs[merc_key]
                ws_resumen.cell(row_cursor, 1, merc_label).font = openpyxl.styles.Font(bold=True)
                row_cursor += 1
                headers = ['Bloque'] + [r['dia'] for r in m['resultados']] + ['Total semana']
                for j, h in enumerate(headers):
                    ws_resumen.cell(row_cursor, 1 + j, h).font = openpyxl.styles.Font(bold=True)
                row_cursor += 1
                bloques_todos = {}
                for r in m['resultados']:
                    for code, (nombre, _s) in r['shares'].items():
                        bloques_todos.setdefault(code, nombre)
                for code, nombre in bloques_todos.items():
                    ws_resumen.cell(row_cursor, 1, nombre)
                    total = 0
                    for j, r in enumerate(m['resultados']):
                        v = r['asignado'].get(code, 0)
                        ws_resumen.cell(row_cursor, 2 + j, v)
                        total += v
                    ws_resumen.cell(row_cursor, 2 + len(m['resultados']), total)
                    row_cursor += 1
                ws_resumen.cell(row_cursor, 1, 'Sobrante sin asignar fin de semana')
                ws_resumen.cell(row_cursor, 2, len(m['sobrante_fin_semana']))
                row_cursor += 2

            for merc_key, merc_label in [('capon', 'Capón'), ('chancha', 'Chancha')]:
                m = rs[merc_key]
                for r in m['resultados']:
                    dia_corto = r['dia']
                    ws_r = wb.create_sheet(f'{dia_corto} Resumen {merc_label[:3]}')
                    write_resumen_dia(ws_r, dia_corto, {
                        merc_label: (r['shares'], r['presupuesto'], r['asignado'], r['disponible_dia'], r['total_asignado_dia']),
                    })
                    ws_c = wb.create_sheet(f'{dia_corto} {merc_label} Correl')
                    write_correlativos(ws_c, merc_label, r['correlativos'], r['shares'])
                    ws_s = wb.create_sheet(f'{dia_corto} {merc_label} Sobr')
                    write_sobrante(ws_s, merc_label, r['sobrante'])

            buf_semana = io.BytesIO()
            wb.save(buf_semana)
            st.download_button('Descargar Excel de la semana', buf_semana.getvalue(),
                                file_name=f"Propuesta reparto semanal - Semana {rs['semana_iso']} - {datetime.date.today().isoformat()}.xlsx",
                                key='descargar_excel_semana')

            if st.button('Guardar toda la semana'):
                for merc_key, merc_label in [('capon', 'Capón'), ('chancha', 'Chancha')]:
                    m = rs[merc_key]
                    for r in m['resultados']:
                        movidos = {mv['correlativo'] for mv in m['movimientos_golpes_cortes_por_dia'].get(r['dia'], []) if mv['destino']}
                        filas_guardar = []
                        for code, (_nombre, _s) in r['shares'].items():
                            for a in r['correlativos'].get(code, []):
                                filas_guardar.append({
                                    'mercaderia': merc_label, 'bloque_codigo': code, 'correlativo': a['correlativo'],
                                    'kg': a['kg'], 'nivel_grasa': a['tipif'], 'observacion': a.get('observacion'),
                                    'reasignado_por_golpe_corte': a['correlativo'] in movidos,
                                })
                        if filas_guardar:
                            db.insert_reparto_resultados(r['dia'], filas_guardar)
                st.success('Semana completa guardada.')


# ============================== ACTUALIZAR DATOS ==============================
elif pagina == 'Actualizar datos':
    st.title('Actualizar datos desde Excel')
    st.caption('Subí los mismos archivos que ya usás — no hace falta tocar nada de ellos.')

    st.subheader('Stock (STOCK (BD) del Excel de Dropbox)')
    st.caption('Copiá el rango de filas de la hoja "STOCK (BD)" (columnas A a N completas, sin '
               'encabezado) y pegalo acá con Ctrl+V — mismo orden de columnas que el Excel, así '
               'entra tal cual. Reemplaza el stock anterior entero.')

    # Estado actual guardado en la base — Tomás, 2026-08-13: "le di refresh y desapareció el
    # stock que ya había cargado". La tabla de pegado SIEMPRE arranca vacía (lista para el
    # próximo pegado) después de guardar o de refrescar la página — no significa que se haya
    # perdido nada, lo guardado ya está en la base. Esto lo confirma en pantalla.
    _snap_info = db.fetch_stock_snapshot_info()
    if _snap_info['ultimo_sync']:
        _ts = datetime.datetime.fromisoformat(_snap_info['ultimo_sync'].replace('Z', '+00:00'))
        st.info(f"Stock guardado actualmente: **{_snap_info['total_filas']} piezas** "
                f"(cargado el {_ts.astimezone().strftime('%d/%m %H:%M')}). La tabla de abajo "
                f"arranca vacía siempre — es para el próximo pegado, no borra lo ya guardado.")

    cols_stock = ['Productor', 'Fecha faena', 'Tipificación OP', 'Correlativo', 'Kg', 'X',
                  'Mercadería', 'Conf.', 'Gras.', 'Garrón', 'Tropa', 'Cat', 'Calidad', 'Observaciones']

    def _df_stock_vacio():
        # dtype='object' explícito — una lista vacía [] cae en float64 por defecto en pandas,
        # lo que choca con la columna "Fecha faena" configurada como texto (Tomás, 2026-08-13,
        # StreamlitAPIException: "column type text... not compatible... ColumnDataKind.FLOAT").
        # UNA fila vacía (no cero) — Tomás, 2026-08-13: "copio y pego del excel y lo pega así y
        # queda mal" (todo el bloque pegado caía en una sola celda). Con la grilla en 0 filas no
        # hay ninguna celda concreta donde anclar el pegado multi-fila/columna y el data_editor
        # de Streamlit lo tira todo como texto suelto en la primera celda; con 1 fila de arranque
        # sí hay una celda (0,0) para pararse y pegar, y de ahí crece solo (num_rows='dynamic').
        return pd.DataFrame({c: [''] for c in cols_stock})

    if 'df_stock_pegado' not in st.session_state:
        st.session_state['df_stock_pegado'] = _df_stock_vacio()
    # Sin column_config con tipo forzado para "Fecha faena": Streamlit vuelve a inferir el
    # dtype de cada columna a partir de lo pegado (Ctrl+V) en cada rerun, y si ese pegado deja
    # la columna vacía/numérica en algún momento, un TextColumn forzado choca contra eso y
    # tira StreamlitAPIException (Tomás, 2026-08-13, se repitió incluso con el seed en
    # dtype='object'). parse_stock_pegado() ya soporta texto o fecha en esa columna.
    df_stock_editado = st.data_editor(
        st.session_state['df_stock_pegado'], key='editor_stock', num_rows='dynamic',
        use_container_width=True, height=350,
    )
    if st.button('Cargar stock desde la tabla', type='primary'):
        try:
            rows = excel_import.parse_stock_pegado(df_stock_editado)
            if not rows:
                st.error('No encontré filas válidas (necesitan Correlativo y Kg) — ¿pegaste el rango completo?')
            else:
                db.replace_stock_snapshot(rows)
                # Tomás, 2026-08-13: "olvidate los datos anteriores y las asignaciones
                # anteriores, empecemos de cero... resetea el historial y tomalo como el stock
                # disponible total" — cada pegado de stock arranca de cero, sin arrastrar
                # repartos guardados de corridas viejas.
                db.reset_historial_reparto()
                st.session_state.pop('resultado', None)
                st.session_state.pop('resultado_semana', None)
                st.success(f'Cargado — {len(rows)} piezas de stock. Historial de repartos reseteado: '
                           f'este stock es el 100% disponible.')
                st.session_state['df_stock_pegado'] = _df_stock_vacio()
                st.rerun()
        except Exception as e:
            st.error(f'Error al leer la tabla: {e}')

    with st.expander('Prefiero subir el archivo Excel completo'):
        f_stock = st.file_uploader('Excel de stock', type=['xlsx'], key='up_stock')
        if f_stock is not None and st.button('Cargar stock desde el Excel'):
            try:
                rows = excel_import.parse_stock_bd(f_stock)
                db.replace_stock_snapshot(rows)
                db.reset_historial_reparto()
                st.session_state.pop('resultado', None)
                st.session_state.pop('resultado_semana', None)
                st.success(f'Cargado — {len(rows)} piezas de stock. Historial de repartos reseteado.')
            except Exception as e:
                st.error(f'Error al leer el Excel: {e}')

    st.divider()
    st.subheader('Cupo semanal por bloque (cantidades)')
    st.caption('Editá acá directamente las cantidades — no hace falta preparar un Excel. La '
               'calidad (tipificación/peso) asignada a cada bloque sigue saliendo del promedio '
               'histórico de abajo, esto no la toca.')
    tab_capon_cupo, tab_chancha_cupo = st.tabs(['Artículo: Capones', 'Artículo: Chanchas'])
    for merc_cupo, tab_cupo in [('Capón', tab_capon_cupo), ('Chancha', tab_chancha_cupo)]:
        with tab_cupo:
            filas_actuales = db.fetch_cupo_bloque(merc_cupo)
            por_bloque = {}
            for f in filas_actuales:
                por_bloque.setdefault(f['bloque_codigo'], {'nombre': f['bloque_nombre'], 'dias': {}})
                por_bloque[f['bloque_codigo']]['dias'][f['dia']] = f['cupo']
            data = []
            for code, d in sorted(por_bloque.items()):
                fila = {'Bloque': d['nombre']}
                for dia in DIAS_HABILES:
                    fila[dia] = d['dias'].get(dia, 0.0)
                data.append(fila)
            df_cupo = pd.DataFrame(data, columns=['Bloque'] + DIAS_HABILES)
            df_editado = st.data_editor(
                df_cupo, key=f'editor_cupo_{merc_cupo}', num_rows='dynamic', use_container_width=True,
                column_config={dia: st.column_config.NumberColumn(dia, min_value=0, step=1) for dia in DIAS_HABILES},
            )

            # Totales calculados en vivo a partir de lo que se acaba de editar (no de lo
            # guardado) — Tomás, 2026-08-12: "quiero que los totales... sumen automáticamente".
            filas_validas = df_editado[df_editado['Bloque'].astype(str).str.strip() != '']
            cols_total = st.columns(len(DIAS_HABILES) + 1)
            for i, dia in enumerate(DIAS_HABILES):
                cols_total[i].metric(dia, int(filas_validas[dia].fillna(0).sum()))
            cols_total[-1].metric('Total semana', int(filas_validas[DIAS_HABILES].fillna(0).sum().sum()))
            with st.expander('Total semanal por bloque'):
                st.dataframe(
                    pd.DataFrame({'Bloque': filas_validas['Bloque'],
                                  'Total semanal': filas_validas[DIAS_HABILES].fillna(0).sum(axis=1)}),
                    use_container_width=True, hide_index=True,
                )

            if st.button(f'Guardar cupos — {merc_cupo}', key=f'guardar_cupo_{merc_cupo}'):
                filas_guardar = []
                for _, fila in df_editado.iterrows():
                    bloque_txt = str(fila['Bloque']).strip()
                    if not bloque_txt or bloque_txt.upper() == 'TOTAL':
                        continue
                    codigo = bloque_txt.split(' - ', 1)[0].strip()
                    dias_bloque = {dia: float(fila[dia] or 0) for dia in DIAS_HABILES}
                    dias_bloque['Total'] = sum(dias_bloque.values())
                    for dia, cupo in dias_bloque.items():
                        filas_guardar.append({'bloque_codigo': codigo, 'bloque_nombre': bloque_txt,
                                               'dia': dia, 'cupo': cupo})
                db.update_cupos_bloque(merc_cupo, filas_guardar)
                st.success(f'Cupos de {merc_cupo} guardados.')
                st.rerun()

    st.divider()
    st.subheader('Histórico de bloques (Promedio diario Capón y Chancha por vendedor.xlsx)')
    st.caption('% de peso/tipificación por bloque (calidad) — y también reemplaza el cupo de '
               'cantidades de arriba entero si subís este archivo. Usalo solo si querés recargar '
               'todo de una desde el Excel; para ajustar cantidades sueltas, usá la grilla de arriba.')
    f_hist = st.file_uploader('Excel histórico', type=['xlsx'], key='up_hist')
    if f_hist is not None and st.button('Cargar histórico'):
        try:
            bloque_rows, calidad_rows = excel_import.parse_historico(f_hist)
            db.replace_historico(bloque_rows, calidad_rows)
            st.success(f'Cargado — {len(bloque_rows)} filas de cupo, {len(calidad_rows)} filas de calidad.')
        except Exception as e:
            st.error(f'Error al leer el Excel: {e}')
